"""
RenpyBox - Excel 导入/导出处理模块
支持翻译结果的 Excel 格式导入和导出

Author: RenpyBox Team
Date: 2025-10-20
Version: 0.1.0
"""

import os
import math
from pathlib import Path
from typing import Dict, List, Optional, TYPE_CHECKING
from datetime import datetime

from base.Version import Version

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

TRANSLATION_HEADERS = ["文件", "行号", "原文", "译文", "类型", "状态", "备注"]

from base.LogManager import LogManager

logger = LogManager.get()


class ExcelExporter:
    """Excel 导出器"""

    def __init__(self):
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas 未安装，请运行: pip install pandas")

    def export(
        self,
        translations: Dict[str, List[Dict]],
        output_path: str,
        include_metadata: bool = True,
        per_file_sheet: bool = False,
    ) -> bool:
        """导出翻译为 Excel 格式"""

        try:
            logger.info(f"开始导出 Excel: {output_path}")

            rows = self._build_rows(translations)
            df = pd.DataFrame(rows, columns = TRANSLATION_HEADERS)

            engine = "openpyxl" if EXCEL_AVAILABLE else None
            with pd.ExcelWriter(output_path, engine = engine) as writer:
                if include_metadata:
                    meta_df = pd.DataFrame(self._metadata_rows(translations), columns = ["键", "值"])
                    meta_df.to_excel(writer, sheet_name = "元数据", index = False)

                if per_file_sheet and not df.empty:
                    for sheet_name, group in self._iter_group_by_file(df):
                        group.to_excel(writer, sheet_name = sheet_name, index = False)
                else:
                    df.to_excel(writer, sheet_name = "翻译", index = False)

            if EXCEL_AVAILABLE:
                self._apply_styles(output_path, include_metadata, per_file_sheet)

            logger.info(f"Excel 导出成功: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Excel 导出失败: {str(e)}", e)
            return False

    def _build_rows(self, translations: Dict[str, List[Dict]]) -> List[Dict]:
        rows: List[Dict] = []
        for file_path, items in translations.items():
            for item in items:
                rows.append(
                    {
                        "文件": str(file_path),
                        "行号": item.get("line", ""),
                        "原文": item.get("original", ""),
                        "译文": item.get("translation", ""),
                        "类型": item.get("type", "dialogue"),
                        "状态": item.get("status", "pending"),
                        "备注": item.get("note", ""),
                    }
                )
        return rows

    def _iter_group_by_file(self, df):
        counts: Dict[str, int] = {}
        for file_name, group in df.groupby("文件", dropna = False):
            sheet_name = self._resolve_sheet_name(str(file_name or "sheet"), counts)
            yield sheet_name, group

    def _resolve_sheet_name(self, name: str, counts: Dict[str, int]) -> str:
        base = (Path(name).stem or "sheet")[:31]
        if base not in counts:
            counts[base] = 0
            return base

        counts[base] += 1
        suffix = counts[base]
        return f"{base[:28]}_{suffix}"

    def _apply_styles(self, output_path: str, include_metadata: bool, per_file_sheet: bool) -> None:
        wb = openpyxl.load_workbook(output_path)

        if include_metadata and "元数据" in wb.sheetnames:
            self._style_metadata_sheet(wb["元数据"])

        target_sheets: List[str]
        if per_file_sheet:
            target_sheets = [name for name in wb.sheetnames if name not in ("元数据", "Metadata")]
        else:
            target_sheets = [name for name in ("翻译",) if name in wb.sheetnames]

        for sheet_name in target_sheets:
            self._style_translation_sheet(wb[sheet_name])

        wb.save(output_path)

    def _style_translation_sheet(self, ws: "Worksheet"):
        """统一应用翻译工作表样式"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(TRANSLATION_HEADERS) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        widths = [40, 8, 50, 50, 12, 12, 30]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A2'
        last_row = max(ws.max_row, 1)
        end_col = get_column_letter(len(TRANSLATION_HEADERS))
        ws.auto_filter.ref = f"A1:{end_col}{last_row}"

    def _style_metadata_sheet(self, ws: "Worksheet"):
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _metadata_rows(self, translations: Dict) -> List[List[str]]:
        return [
            ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["文件总数", str(len(translations))],
            ["翻译总数", str(sum(len(items) for items in translations.values()))],
            ["工具版本", f"RenpyBox {Version.CURRENT}"],
        ]


class ExcelImporter:
    """Excel 导入器"""
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")
    
    def import_translations(self, excel_path: str) -> Dict[str, List[Dict]]:
        """从 Excel 导入翻译
        
        Args:
            excel_path: Excel 文件路径
            
        Returns:
            翻译数据字典，格式同 ExcelExporter.export()
        """
        try:
            logger.info(f"开始导入 Excel: {excel_path}")

            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"文件不存在: {excel_path}")

            if PANDAS_AVAILABLE:
                try:
                    translations = self._import_with_pandas(excel_path)
                    if translations:
                        logger.info(f"Excel 导入成功: 共 {len(translations)} 个文件 (pandas)")
                        return translations
                except Exception as pandas_error:
                    logger.warning(f"pandas 解析失败，回退 openpyxl: {pandas_error}")
            else:
                logger.info("pandas 未安装，使用 openpyxl 导入（速度可能较慢）")

            translations = self._import_with_openpyxl(excel_path)
            logger.info(f"Excel 导入成功: 共 {len(translations)} 个文件")
            return translations

        except Exception as e:
            logger.error(f"Excel 导入失败: {str(e)}", e)
            return {}
    
    def apply_translations(self, 
                          translations: Dict[str, List[Dict]], 
                          game_dir: str,
                          target_language: str = 'chinese',
                          backup: bool = True) -> bool:
        """将导入的翻译应用到游戏文件
        
        Args:
            translations: 翻译数据字典
            game_dir: 游戏目录（包含 game/ 文件夹）
            target_language: 目标语言
            backup: 是否备份原文件
            
        Returns:
            bool: 是否成功
        """
        try:
            logger.info("开始应用翻译到游戏文件")
            
            game_path = Path(game_dir)
            if not game_path.exists():
                raise FileNotFoundError(f"游戏目录不存在: {game_dir}")
            
            # 创建翻译目录
            tl_dir = game_path / 'game' / 'tl' / target_language
            tl_dir.mkdir(parents=True, exist_ok=True)
            
            seen_global: set[str] = set()
            dup_count = 0

            # 应用每个文件的翻译
            for file_name, items in translations.items():
                # 创建翻译文件路径
                trans_file = tl_dir / file_name
                
                # 备份已有文件
                if backup and trans_file.exists():
                    backup_file = trans_file.with_suffix('.rpy.bak')
                    trans_file.rename(backup_file)
                    logger.info(f"已备份: {backup_file}")
                
                # 生成翻译文件内容
                # 跨文件去重：同一个 old 只保留首次，避免 Ren'Py 重复翻译报错
                filtered_items: List[Dict] = []
                for item in items:
                    original_raw = item.get("original") or ""
                    if original_raw in seen_global:
                        dup_count += 1
                        continue
                    seen_global.add(original_raw)
                    filtered_items.append(item)

                content = self._generate_translation_file(filtered_items, target_language)
                
                # 写入文件
                with open(trans_file, 'w', encoding='utf-8') as f:
                    f.write(content)
                
                logger.info(f"已应用翻译: {trans_file}")
            
            if dup_count > 0:
                logger.warning(f"跳过 {dup_count} 条重复原文（跨文件仅保留首次 old），避免 Ren'Py 重复翻译报错")

            logger.info("翻译应用完成")
            return True
            
        except Exception as e:
            logger.error(f"应用翻译失败: {str(e)}", e)
            return False

    def _import_with_pandas(self, excel_path: str) -> Dict[str, List[Dict]]:
        frames = pd.read_excel(excel_path, sheet_name=None)
        translations: Dict[str, List[Dict]] = {}

        for sheet_name, df in frames.items():
            if sheet_name in ['元数据', 'Metadata']:
                continue
            if df is None or df.empty:
                continue

            df = df.fillna("")
            if not {'原文', '译文'}.issubset(set(df.columns)):
                logger.warning(f"Sheet '{sheet_name}' 缺少必需列，跳过")
                continue

            for _, row in df.iterrows():
                original = str(row.get('原文', '')).strip()
                if not original:
                    continue

                translation = str(row.get('译文', '')).strip()
                file_name = row.get('文件') or f"{sheet_name}.rpy"
                file_name = str(file_name).strip()
                if not file_name.endswith('.rpy') and '/' not in file_name and '\\' not in file_name:
                    file_name = f"{file_name}.rpy"

                item = {
                    'line': self._safe_int(row.get('行号')),
                    'original': original,
                    'translation': translation,
                    'type': str(row.get('类型', 'dialogue')) or 'dialogue',
                    'status': str(row.get('状态', 'translated')) or 'translated',
                }
                translations.setdefault(file_name, []).append(item)

        return translations

    def _import_with_openpyxl(self, excel_path: str) -> Dict[str, List[Dict]]:
        wb = openpyxl.load_workbook(excel_path)
        translations: Dict[str, List[Dict]] = {}

        for sheet_name in wb.sheetnames:
            if sheet_name in ['元数据', 'Metadata']:
                continue

            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue

            headers = [cell.value for cell in ws[1]]
            if not headers:
                continue

            required_cols = ['原文', '译文']
            if not all(col in headers for col in required_cols):
                logger.warning(f"Sheet '{sheet_name}' 缺少必需列，跳过")
                continue

            col_indices = {header: idx for idx, header in enumerate(headers)}
            file_col = col_indices.get('文件')
            line_col = col_indices.get('行号')
            original_col = col_indices.get('原文')
            translation_col = col_indices.get('译文')
            type_col = col_indices.get('类型')
            status_col = col_indices.get('状态')

            for row in ws.iter_rows(min_row=2, values_only=True):
                original = row[original_col] if original_col is not None else None
                if not original:
                    continue

                translation = row[translation_col] if translation_col is not None else ''
                if file_col is not None and row[file_col]:
                    file_name = row[file_col]
                else:
                    file_name = sheet_name + '.rpy'
                file_name = str(file_name)

                item = {
                    'line': self._safe_int(row[line_col]) if line_col is not None else 0,
                    'original': original,
                    'translation': translation,
                    'type': row[type_col] if type_col is not None and row[type_col] else 'dialogue',
                    'status': row[status_col] if status_col is not None and row[status_col] else 'translated',
                }
                translations.setdefault(file_name, []).append(item)

        return translations

    def _safe_int(self, value) -> int:
        if value in (None, ""):
            return 0
        try:
            if isinstance(value, float) and math.isnan(value):
                return 0
            return int(float(value))
        except (ValueError, TypeError):
            return 0
    
    def _generate_translation_file(self, items: List[Dict], language: str) -> str:
        """生成 Ren'Py 翻译文件内容"""
        lines = [
            "# TODO: Translation updated at " + datetime.now().strftime("%Y-%m-%d %H:%M"),
            "# Imported from Excel",
            "",
            f"translate {language} strings:",
            ""
        ]
        
        for item in items:
            original = item['original']
            translation = item.get('translation', original)
            
            # 转义引号
            original = original.replace('"', '\\"')
            translation = translation.replace('"', '\\"')
            
            lines.append(f'    # Line {item.get("line", "?")}')
            lines.append(f'    old "{original}"')
            lines.append(f'    new "{translation}"')
            lines.append('')
        
        return '\n'.join(lines)


def check_openpyxl_installed() -> bool:
    """检查 openpyxl 是否已安装"""
    return EXCEL_AVAILABLE


def install_openpyxl():
    """安装 openpyxl（需要管理员权限）"""
    import subprocess
    import sys
    
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        logger.info("openpyxl 安装成功")
        return True
    except Exception as e:
        logger.error(f"openpyxl 安装失败: {str(e)}", e)
        return False
