# -*- coding: utf-8 -*-
"""Emoji 替换工具：生成译前/译后替换表，执行批量替换。"""

from __future__ import annotations

import re
import os
import shutil
import time
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 嵌入的 Emoji 库（与原脚本一致）
GRS_EMBEDDED_EMOJIS: List[str] = [
    "🌳", "🌴", "🌵", "🌷", "🌹", "🌺", "🌻", "🌼", "🌾", "🌿", "🍀",
    "🍁", "🍂", "🍃", "🍇", "🍉", "🍊", "🍋", "🍌", "🍎", "🍔",
    "🍕", "🍟", "🍦", "🍰", "🍷", "🍹", "🍺", "🎀", "🎁", "🎂",
    "🎈", "🎉", "🎊", "🎓", "🎙️", "🎤", "🎧", "🎨", "🎩", "🎭",
    "🎮", "🎯", "🎰", "🎱", "🎲", "🎳", "🎵", "🎶", "🎸", "🎹",
    "🎺", "🎻", "🎼", "🎽", "🎾", "🎿", "🏀", "🏁", "🏂", "🏃",
    "🏄", "🏅", "🏆", "🏇", "🏈", "🏉", "🏊", "🏋️", "🏐", "🏠",
    "🏡", "🏢", "🏥", "🏦", "🏪", "🏫", "🏬", "🏰", "🐱", "🐶",
    "👦", "👧", "👨", "👩", "👮", "👱", "👲", "👳", "👴", "👵",
    "👶", "👷", "👹", "👺", "👻", "👽", "👿", "💀", "💂", "💓",
    "💔", "💕", "💗", "💘", "💙", "💚", "💛", "💜", "💞", "🕴️",
    "🕵", "🗻", "🗼", "😁", "😂", "😃", "😄", "😅", "😆", "😇",
    "😈", "😉", "😊", "😋", "😌", "😍", "😎", "😏", "😒", "😓",
    "😔", "😖", "😗", "😘", "😙", "😚", "😛", "😜", "😝", "😞",
    "😟", "😠", "😡", "😢", "😥", "😨", "😪", "😰", "😱", "😳",
    "😴", "😵", "😷", "😸", "😹", "😺", "😻", "😼", "😽", "😾",
    "😿", "🙀", "🙂", "🙈", "🙉", "🙊", "🚣", "🚴", "🚵", "🚶",
    "🛀", "🛋️", "🛌", "🛍️", "🛎️", "🛏️", "🛒", "🤐", "🤑", "🤒",
    "🤔", "🤕", "🤖", "🤗", "🤢", "🤣", "🤥", "🤧", "🤩", "🤪",
    "🤫", "🤭", "🤮", "🤯", "🥅", "🥰", "🥵", "🥶", "🧓", "🧕",
    "📱", "📲", "☎", "📞", "📟", "📠", "🔋", "🔌", "💻", "💽",
    "💾", "💿", "📀", "🎥", "📺", "📷", "📹", "📼", "🔍", "🔎",
    "🔬", "🔭", "📡", "📔", "📕", "📖", "📗", "📘", "📙", "📚",
    "📓", "📃", "📜", "📄", "📰", "📑", "🔖", "💳", "✉", "📧",
    "📨", "📩", "📤", "📥", "📦", "📫", "📪", "📬", "📭", "📮",
    "✏", "✒", "📝", "📁", "📂", "📅", "📆", "📇", "📈", "📉",
    "📊", "📋", "📌", "📍", "📎", "📏", "📐", "✂", "🔒", "🔓",
    "🔏", "🔐", "🔑", "🚂", "🚃", "🚄", "🚅", "🚆", "🚇", "🚈",
    "🚉", "🚊", "🚝", "🚞", "🚋", "🚌", "🚍", "🚎", "🚏", "🚐",
    "🚑", "🚒", "🚓", "🚔", "🚕", "🚖", "🚗", "🚘", "🚚", "🚛",
    "🚜", "🚲", "⛽", "🚨", "🚥", "🚦", "🚧", "⚓", "⛵", "🚤",
    "🚢", "✈", "💺", "🚁", "🚟", "🚠", "🚡", "🚀", "🌋", "🏣",
    "🏤", "🏨", "🏩", "🏭", "🏯", "💒", "🗽", "⛪", "⛲", "🌃",
    "🎠", "🎡", "🎢", "🎑", "🗿", "🛂", "🛃", "🛄", "🛅",
]


def _extract_bracketed_text(file_path: Path) -> Dict[str, set[str]]:
    patterns = {
        "{}": re.compile(r"(\{[^{}]*\})"),
        "[]": re.compile(r"((?<!\[)\[[^[\]]*\])"),
    }
    found = {"{}": set(), "[]": set()}
    content: str
    try:
        content = file_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return found

    for bracket, pattern in patterns.items():
        matches = pattern.findall(content)
        found[bracket].update(m for m in matches if len(m) > 2)
    return found


def _sort_key(text: str) -> Tuple[int, int, str]:
    bracket_type = 0 if text.startswith("{") else 1
    length = len(text)
    return (bracket_type, -length, text)


def _generate_combinations(emoji_list: Iterable[str], count: int) -> List[str] | None:
    emojis = list(emoji_list)
    combos: List[str] = []
    total_possible = len(emojis) * (len(emojis) - 1) // 2
    if count > total_possible:
        return None
    for i in range(len(emojis)):
        for j in range(i + 1, len(emojis)):
            combos.append(emojis[i] + emojis[j])
            if len(combos) >= count:
                return combos
    return combos


def generate_emoji_replacement_sheets(tl_dir: Path, output_dir: Path) -> Tuple[int, Path, Path]:
    """扫描 tl 目录并生成译前/译后替换表，返回 (条目数, 译前路径, 译后路径)。"""
    if not tl_dir.exists():
        raise FileNotFoundError(f"语言目录不存在: {tl_dir}")

    rpy_files = list(tl_dir.rglob("*.rpy"))
    if not rpy_files:
        raise ValueError("未找到 .rpy 文件")

    all_text = {"{}": set(), "[]": set()}
    for rpy in rpy_files:
        result = _extract_bracketed_text(rpy)
        all_text["{}"].update(result["{}"])
        all_text["[]"].update(result["[]"])

    combined_text = sorted(list(all_text["{}"]) + list(all_text["[]"]), key=_sort_key)
    total = len(combined_text)

    if total == 0:
        raise ValueError("未找到特殊括号文本")
    if total > 71631:
        raise ValueError("特殊括号文本数量过多(>71631)，无法生成足够的 Emoji 组合")

    emoji_list = GRS_EMBEDDED_EMOJIS
    if total <= len(emoji_list):
        replacements = emoji_list[:total]
    else:
        replacements = _generate_combinations(emoji_list, total)
        if not replacements or len(replacements) < total:
            raise RuntimeError("无法生成足够的唯一 Emoji 组合")

    output_dir.mkdir(parents=True, exist_ok=True)
    pre_path = output_dir / "译前替换.xlsx"
    post_path = output_dir / "译后替换.xlsx"

    # 译前
    wb_pre = Workbook()
    ws_pre = wb_pre.active
    ws_pre.title = "译前替换"
    for text, emoji in zip(combined_text, replacements):
        ws_pre.append([text, emoji])
    _auto_width(ws_pre, 2)
    wb_pre.save(pre_path)

    # 译后
    wb_post = Workbook()
    ws_post = wb_post.active
    ws_post.title = "译后替换"
    for text, emoji in zip(combined_text, replacements):
        ws_post.append([emoji, text])
    _auto_width(ws_post, 2)
    wb_post.save(post_path)

    return total, pre_path, post_path


def _auto_width(sheet, cols: int):
    for col in range(1, cols + 1):
        max_len = 0
        for cell in sheet[get_column_letter(col)]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(col)].width = (max_len + 2) * 1.2


def load_replacement_map(xlsx_path: Path) -> Dict[str, str]:
    """从 Excel 加载替换映射表。"""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"未找到替换表: {xlsx_path}")
    df = pd.read_excel(xlsx_path, header=None)
    if df.shape[1] < 2:
        raise ValueError("替换表至少需要两列数据")
    mapping: Dict[str, str] = {}
    for _, row in df.iterrows():
        key = str(row[0]).strip()
        value = str(row[1]).strip()
        if key and value and key != "nan" and value != "nan":
            mapping[key] = value
    return mapping


def apply_replacements(src_path: Path, mapping: Dict[str, str], dest_path: Path) -> int:
    """执行批量替换，返回匹配到的 key 数量。"""
    content = src_path.read_text(encoding="utf-8")
    sorted_keys = sorted(mapping.keys(), key=lambda x: len(x), reverse=True)
    matched = 0
    for key in sorted_keys:
        if key in content:
            matched += 1
            content = content.replace(key, mapping[key])
    dest_path.write_text(content, encoding="utf-8")
    return matched


# -------------------- 便捷入口（批量版） -------------------- #
def find_mapping_path(base_dir: Path, filename: str) -> Path | None:
    """优先查找 v7 哈基米路径，再兼容 v6 路径。"""
    path_v7 = base_dir / "translate_output" / "3_Emoji_Tools" / filename
    path_v6 = base_dir / "translate" / "4_Emoji替换表" / filename
    if path_v7.exists():
        return path_v7
    if path_v6.exists():
        return path_v6
    return None


def load_default_mapping(base_dir: Path, mode: str) -> Dict[str, str]:
    """根据模式加载默认映射表。mode=prepare|restore"""
    filename = "Tag_Protection_Pre(译前).xlsx" if mode == "prepare" else "Tag_Protection_Post(译后).xlsx"
    alt = "译前替换.xlsx" if mode == "prepare" else "译后替换.xlsx"
    path = find_mapping_path(base_dir, filename) or find_mapping_path(base_dir, alt)
    if not path:
        raise FileNotFoundError(f"未找到默认映射表，请先生成 Emoji/Tag 对照表 ({filename}/{alt})")
    return load_replacement_map(path)


def backup_folder(src_folder: Path) -> Path:
    """拷贝备份目录，返回备份路径；若存在则自动生成唯一目录。"""
    dir_name = src_folder.parent
    base_name = src_folder.name
    timestamp = int(time.time())
    backup_path = dir_name / f"{base_name}_backup_{timestamp}"
    counter = 1
    while backup_path.exists():
        backup_path = dir_name / f"{base_name}_backup_{timestamp}_{counter}"
        counter += 1
    shutil.copytree(src_folder, backup_path)
    return backup_path


def apply_replacements_dir(folder: Path, mapping: Dict[str, str], *, is_restore: bool = False) -> Tuple[int, int]:
    """对目录下所有 .rpy 进行替换，返回 (成功文件数, 失败数)。"""
    success = 0
    failed = 0
    for root, _, files in os.walk(folder):
        for file in files:
            if not file.endswith(".rpy"):
                continue
            file_path = Path(root) / file
            try:
                content = file_path.read_text(encoding="utf-8")
                original = content
                for key, val in mapping.items():
                    if key in content:
                        content = content.replace(key, val)
                if content != original:
                    file_path.write_text(content, encoding="utf-8")
                success += 1
            except Exception:
                failed += 1
    return success, failed


__all__ = [
    "GRS_EMBEDDED_EMOJIS",
    "generate_emoji_replacement_sheets",
    "load_replacement_map",
    "apply_replacements",
]
