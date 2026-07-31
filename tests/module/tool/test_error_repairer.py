from importlib import import_module
from types import SimpleNamespace

import pytest

from module.Tool.ErrorRepairer import ErrorRepairer


error_repairer_module = import_module("module.Tool.ErrorRepairer")


def _quote_errors(repairer: ErrorRepairer, tmp_path, content: str):
    script = tmp_path / "script.rpy"
    script.write_text(content, encoding="utf-8")
    return repairer.check_file(
        str(script),
        check_syntax=False,
        check_indent=False,
        check_indent_level=False,
        check_quotes=True,
        check_dialogue_quotes=True,
    )


def _translation_errors(repairer: ErrorRepairer, tmp_path, content: str):
    script = tmp_path / "translation.rpy"
    script.write_text(content, encoding="utf-8")
    errors = repairer.check_file(
        str(script),
        check_syntax=False,
        check_indent=False,
        check_indent_level=False,
        check_quotes=False,
        check_dialogue_quotes=False,
        check_translation_issues=True,
    )
    return script, errors


def test_repairs_only_inner_unescaped_quotes_with_exact_output() -> None:
    repairer = ErrorRepairer()

    repaired, changed = repairer.repair_unescaped_dialogue_quotes(
        'new "他说"你好"，然后走了"\r\n'
    )

    assert changed is True
    assert repaired == 'new "他说\\"你好\\"，然后走了"\r\n'


def test_repair_preserves_spacing_and_does_not_touch_quoted_comment() -> None:
    repaired, changed = ErrorRepairer().repair_unescaped_dialogue_quotes(
        '    e\t"He said "hello" today"  # keep "comment"\n'
    )

    assert changed is True
    assert repaired == '    e\t"He said \\"hello\\" today"  # keep "comment"\n'


def test_preserves_curly_quotes_inside_a_normal_string(tmp_path) -> None:
    script = tmp_path / "strings.rpy"
    original = 'new "他说“你好”，然后走了"\n'
    script.write_text(original, encoding="utf-8")

    success, fix_count = ErrorRepairer().auto_fix_file(
        str(script),
        fix_indent=False,
        fix_quotes=True,
        fix_dialogue_quotes=True,
    )

    assert success is True
    assert fix_count == 0
    assert script.read_text(encoding="utf-8") == original


def test_normalizes_only_outer_curly_delimiters() -> None:
    repaired, changed = ErrorRepairer().repair_curly_string_delimiters(
        'new “他说“你好”，然后走了” # 保留“注释”\n'
    )

    assert changed is True
    assert repaired == 'new "他说“你好”，然后走了" # 保留“注释”\n'


def test_auto_fix_is_idempotent(tmp_path) -> None:
    script = tmp_path / "strings.rpy"
    script.write_text('old "He said "hello" today"\n', encoding="utf-8")
    repairer = ErrorRepairer()

    first_success, first_count = repairer.auto_fix_file(
        str(script),
        fix_indent=False,
        fix_quotes=True,
        fix_dialogue_quotes=True,
    )
    first_output = script.read_text(encoding="utf-8")
    second_success, second_count = repairer.auto_fix_file(
        str(script),
        fix_indent=False,
        fix_quotes=True,
        fix_dialogue_quotes=True,
    )

    assert first_success is True
    assert first_count == 1
    assert first_output == 'old "He said \\"hello\\" today"\n'
    assert second_success is True
    assert second_count == 0
    assert script.read_text(encoding="utf-8") == first_output


def test_apostrophe_inside_double_quoted_text_is_not_an_error(tmp_path) -> None:
    errors = _quote_errors(ErrorRepairer(), tmp_path, 'old "I don\'t know"\n')

    assert errors == []


@pytest.mark.parametrize(
    "line",
    [
        '$ values = ("one", "two")\n',
        'value = "one" "two"\n',
        'e "one" "two"\n',
        '    "Choice" if persistent.seen_intro:\n',
        '    textbutton "Start" action Return("start")\n',
        'e "Hello" # translator wrote "greeting"\n',
        '    """docstring with "quoted" text"""\n',
        "new 'I don't know'\n",
    ],
)
def test_quote_repair_fails_closed_for_ambiguous_or_legal_lines(line: str) -> None:
    repaired, changed = ErrorRepairer().repair_unescaped_dialogue_quotes(line)

    assert changed is False
    assert repaired == line


def test_scanner_ignores_legal_python_menu_screen_comments_and_triple_quotes(tmp_path) -> None:
    content = (
        '$ values = ("one", "two")\n'
        'value = "one" "two"\n'
        'e "one" "two"\n'
        '    "Choice" if persistent.seen_intro:\n'
        '    textbutton "Start" action Return("start")\n'
        'e "Hello" # translator wrote "greeting"\n'
        '    """docstring with "quoted" text"""\n'
    )

    assert _quote_errors(ErrorRepairer(), tmp_path, content) == []


def test_scanner_reports_repairable_and_unterminated_translation_strings(tmp_path) -> None:
    errors = _quote_errors(
        ErrorRepairer(),
        tmp_path,
        'old "He said "hello" today"\nnew "unterminated\n',
    )

    assert [(error["line"], error["type"]) for error in errors] == [
        (1, "quotes"),
        (2, "quotes"),
    ]


def test_placeholder_scan_uses_counter_and_distinguishes_issue_types(tmp_path) -> None:
    _, errors = _translation_errors(
        ErrorRepairer(),
        tmp_path,
        '''translate schinese strings:
    old "Hi [name] [name]"
    new "你好 [name]"

    old "Score [score]"
    new "分数 [score] [bonus]"

    old "Welcome [player_name]"
    new "欢迎 [玩家名]"

    old "Literal [[player_name]] and [kept]"
    new "字面 [[玩家名]] 和 [kept]"
''',
    )

    placeholder_errors = [
        error for error in errors if error["type"].startswith("placeholder_")
    ]
    assert [(error["line"], error["type"]) for error in placeholder_errors] == [
        (3, "placeholder_missing"),
        (6, "placeholder_extra"),
        (9, "placeholder_rewritten"),
    ]
    assert placeholder_errors[0]["placeholders"] == {"[name]": 1}
    assert placeholder_errors[1]["placeholders"] == {"[bonus]": 1}
    assert placeholder_errors[2]["rewrites"] == [
        {"source": "[player_name]", "target": "[玩家名]", "count": 1}
    ]


def test_placeholder_and_linebreak_scan_supports_dialogue_translation_blocks(tmp_path) -> None:
    _, errors = _translation_errors(
        ErrorRepairer(),
        tmp_path,
        '''translate schinese start_abcd:
    # e "Hello [name]\\nAgain"
    e "你好 [name]"
''',
    )

    assert [(error["line"], error["type"]) for error in errors] == [
        (3, "linebreak_mismatch"),
    ]
    assert errors[0]["source_count"] == 1
    assert errors[0]["target_count"] == 0


def test_placeholder_scan_supports_single_quoted_old_new_pairs(tmp_path) -> None:
    _, errors = _translation_errors(
        ErrorRepairer(),
        tmp_path,
        "translate schinese strings:\n    old 'Hi [name]'\n    new '你好'\n",
    )

    assert [(error["line"], error["type"]) for error in errors] == [
        (3, "placeholder_missing"),
    ]
    assert errors[0]["placeholders"] == {"[name]": 1}


def test_empty_new_is_legal_and_extra_empty_literal_is_reported(tmp_path) -> None:
    _, errors = _translation_errors(
        ErrorRepairer(),
        tmp_path,
        '''translate schinese strings:
    old "Pending [name]\\n"
    new ""

    old "Done"
    new "完成" ""

translate schinese start_abcd:
    # e "Silent"
    e ""
    # e "Spoken"
    e "说完了" ""
    # e "Condition"
    e "选项" if persistent.route == ""
    # e "Condition extra"
    e "选项" if persistent.route == "" ""
''',
    )

    assert [(error["line"], error["type"]) for error in errors] == [
        (6, "extra_empty_string"),
        (12, "extra_empty_string"),
        (16, "extra_empty_string"),
    ]
    assert all(error["count"] == 1 for error in errors)


def test_scan_reports_same_file_duplicate_without_writing(tmp_path) -> None:
    content = '''translate schinese strings:
    old "Same"
    new "相同"

    old "Same"
    new "再次"
'''
    script, errors = _translation_errors(ErrorRepairer(), tmp_path, content)

    assert [error["type"] for error in errors] == ["duplicate_old_new"]
    assert errors[0]["line"] == 5
    assert errors[0]["first_line"] == 2
    assert script.read_text(encoding="utf-8") == content


def test_folder_scan_reports_cross_file_duplicates_per_language_without_writing(tmp_path) -> None:
    first = tmp_path / "a.rpy"
    second = tmp_path / "b.rpy"
    other_language = tmp_path / "c.rpy"
    first.write_text(
        'translate schinese strings:\n    old "Same"\n    new "甲"\n',
        encoding="utf-8",
    )
    second.write_text(
        'translate schinese strings:\n    old "Same"\n    new "乙"\n',
        encoding="utf-8",
    )
    other_language.write_text(
        'translate french strings:\n    old "Same"\n    new "Pareil"\n',
        encoding="utf-8",
    )
    before = {path: path.read_bytes() for path in (first, second, other_language)}

    report = ErrorRepairer().check_folder(
        str(tmp_path),
        check_syntax=False,
        check_indent=False,
        check_indent_level=False,
        check_quotes=False,
        check_dialogue_quotes=False,
        check_translation_issues=True,
    )

    duplicate = report[str(second)][0]
    assert duplicate["type"] == "duplicate_old_new"
    assert duplicate["first_file"] == str(first)
    assert str(other_language) not in report
    assert {path: path.read_bytes() for path in before} == before


def test_translation_scan_never_writes_the_source_file(tmp_path) -> None:
    content = '''translate schinese strings:
    old "Hi [name]\\n"
    new "你好 [名字]" ""
'''
    script = tmp_path / "translation.rpy"
    script.write_text(content, encoding="utf-8")
    before = script.read_bytes()

    ErrorRepairer().check_file(str(script))

    assert script.read_bytes() == before


def test_export_report_preserves_full_path_and_structured_details(tmp_path) -> None:
    report_path = tmp_path / "report.xlsx"
    source_path = str(tmp_path / "nested" / "translation.rpy")
    report = {
        source_path: [{
            "line": 8,
            "type": "placeholder_missing",
            "message": "译文缺少原文占位符",
            "content": 'new "你好"',
            "placeholders": {"[name]": 2},
            "source_line": 7,
        }],
    }

    ErrorRepairer().export_error_report(report, str(report_path))

    from openpyxl import load_workbook

    sheet = load_workbook(report_path).active
    assert [cell.value for cell in sheet[1]] == [
        "文件", "行号", "错误类型", "错误信息", "内容", "详情",
    ]
    assert sheet.cell(2, 1).value == source_path
    assert '"[name]": 2' in sheet.cell(2, 6).value
    assert '"source_line": 7' in sheet.cell(2, 6).value


def _patch_lint_runtime(monkeypatch, tmp_path, result) -> str:
    game_exe = tmp_path / "sample.exe"
    python_exe = tmp_path / "python.exe"
    game_py = tmp_path / "sample.py"
    for path in (game_exe, python_exe, game_py):
        path.write_text("", encoding="utf-8")

    monkeypatch.setattr(
        error_repairer_module,
        "get_python_path_from_game_path",
        lambda _path: str(python_exe),
    )
    monkeypatch.setattr(
        error_repairer_module,
        "get_py_path",
        lambda _path: str(game_py),
    )
    monkeypatch.setattr(error_repairer_module.subprocess, "run", lambda *args, **kwargs: result)
    return str(game_exe)


def test_exec_lint_distinguishes_clean_run_from_failure(monkeypatch, tmp_path) -> None:
    clean_game = _patch_lint_runtime(
        monkeypatch,
        tmp_path,
        SimpleNamespace(returncode=0, stdout="", stderr=""),
    )
    assert ErrorRepairer().exec_renpy_lint(clean_game) == ""

    failed_game = _patch_lint_runtime(
        monkeypatch,
        tmp_path,
        SimpleNamespace(returncode=1, stdout="", stderr="launcher failed"),
    )
    assert ErrorRepairer().exec_renpy_lint(failed_game) is None


@pytest.mark.parametrize(
    "error_type",
    [
        "",
        "unterminated_string",
        "unknown_statement",
        "expected_statement",
        "empty_block",
        "duplicate_translation",
    ],
)
def test_lint_fixer_unknown_or_destructive_types_never_write(tmp_path, error_type: str) -> None:
    script = tmp_path / "script.rpy"
    original = 'old "Keep me"\nnew "保留我"\n'
    script.write_text(original, encoding="utf-8")

    changed = ErrorRepairer()._fix_single_lint_error(str(script), 1, error_type)

    assert changed is False
    assert script.read_text(encoding="utf-8") == original


def test_fix_by_lint_stops_on_unknown_error_without_writing(monkeypatch, tmp_path) -> None:
    game_dir = tmp_path / "game"
    game_dir.mkdir()
    script = game_dir / "script.rpy"
    original = 'label start:\n    "Keep me"\n'
    script.write_text(original, encoding="utf-8")
    repairer = ErrorRepairer()
    monkeypatch.setattr(
        repairer,
        "exec_renpy_lint",
        lambda _path: '"game/script.rpy", line 2: unknown statement',
    )

    success, fixed = repairer.fix_by_lint(str(tmp_path / "sample.exe"))

    assert success is False
    assert fixed == 0
    assert script.read_text(encoding="utf-8") == original


def test_fix_by_lint_repairs_one_safe_error_then_relints(monkeypatch, tmp_path) -> None:
    game_dir = tmp_path / "game"
    game_dir.mkdir()
    script = game_dir / "script.rpy"
    script.write_text('new "He said "hello" today"\n', encoding="utf-8")
    outputs = iter([
        '"game/script.rpy", line 1: end of line expected',
        "",
    ])
    repairer = ErrorRepairer()
    monkeypatch.setattr(repairer, "exec_renpy_lint", lambda _path: next(outputs))

    success, fixed = repairer.fix_by_lint(str(tmp_path / "sample.exe"))

    assert success is True
    assert fixed == 1
    assert script.read_text(encoding="utf-8") == 'new "He said \\"hello\\" today"\n'


def test_parse_lint_error_preserves_file_paths_with_spaces() -> None:
    [error] = ErrorRepairer().parse_lint_errors(
        'File "C:\\Games\\My Game\\game\\script.rpy", line 7: end of line expected'
    )

    assert error["file"] == "C:\\Games\\My Game\\game\\script.rpy"
    assert error["line"] == 7
    assert error["type"] == "syntax_error"
