"""
批量修正页面 - 完整实现
基于 LinguaGacha 的 BatchCorrectionPage 移植
"""
import ast
import os
import re
import json
import shutil
import threading
from pathlib import Path

import openpyxl
import openpyxl.worksheet.worksheet
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import Qt, QUrl
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLayout, QFileDialog
from qfluentwidgets import (
    PushButton,
    InfoBar,
    FluentIcon,
    SingleDirectionScrollArea,
    TransparentPushButton,
)

from base.Base import Base
from base.LogManager import LogManager
from module.Config import Config
from widget.EmptyCard import EmptyCard
from widget.CommandBarCard import CommandBarCard
from widget.ThemeHelper import mark_toolbox_widget, mark_toolbox_scroll_area


class BatchCorrectionPage(Base, QWidget):
    """批量修正页面 - 完整功能实现"""
    
    # 文件名匹配规则
    FILE_NAME_WHITELIST = re.compile(r"^(结果检查_|result_check_)([^\\/]+)\.json$", flags=re.IGNORECASE)
    FILE_NAME_BLACKLIST = ("result_check_untranslated.json", "结果检查_未翻译.json")

    def __init__(self, object_name: str, parent=None):
        Base.__init__(self)
        QWidget.__init__(self, parent)
        self.setObjectName(object_name)
        mark_toolbox_widget(self)
        
        self.window = parent
        self.input_folder = ""   # 需要用户设置
        self.output_folder = ""  # 需要用户设置
        self.translation_root = ""  # 译文文件所在根目录
        
        self._init_ui()

    def _init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(24, 24, 24, 24)

        # 标题卡片
        title_card = EmptyCard(
            title="批量修正",
            description=(
                "根据翻译完成时生成的结果检查文件中的数据，对可能存在的翻译错误进行批量修正，然后生成修正后的译文文件<br><br>"
                "<b>工作流程：</b><br>"
                "• 从输入文件夹的翻译结果检查文件中提取可能需要修正的数据<br>"
                "• 检查提取出的数据，并根据实际情况对需要修正的条目进行修正<br>"
                "• 将修正后的数据注入译文文件，然后在输出文件夹生成修正后的译文文件"
            ),
            init=None,
        )
        layout.addWidget(title_card)

        # 创建滚动区域
        scroll_area = SingleDirectionScrollArea(orient=Qt.Orientation.Vertical)
        scroll_area.setWidgetResizable(True)
        scroll_area.enableTransparentBackground()
        mark_toolbox_scroll_area(scroll_area)

        scroll_widget = QWidget()
        mark_toolbox_widget(scroll_widget, "toolboxScroll")
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(12)

        # 添加步骤卡片
        scroll_layout.addWidget(self._create_step1_card())
        scroll_layout.addWidget(self._create_step2_card())
        scroll_layout.addStretch(1)
        
        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # 底部命令栏
        self.command_bar_card = CommandBarCard()
        layout.addWidget(self.command_bar_card)
        self.command_bar_card.add_stretch(1)
        
        wiki_btn = TransparentPushButton(FluentIcon.HELP, "帮助文档")
        wiki_btn.clicked.connect(lambda: QDesktopServices.openUrl(
            QUrl("https://github.com/dclef/RenpyBox/wiki")
        ))
        self.command_bar_card.add_widget(wiki_btn)

    def _create_step1_card(self) -> EmptyCard:
        """创建步骤一卡片"""
        def init(widget: EmptyCard) -> None:
            btn = PushButton(FluentIcon.PLAY, "开始")
            btn.clicked.connect(self._step_01_clicked)
            widget.add_widget(btn)

        return EmptyCard(
            title="第一步 - 生成修正数据",
            description=(
                "从结果检查文件中提取可能包含翻译错误的数据<br>"
                "然后自动在输出文件夹内生成用于编辑的数据文件 <b>批量修正.xlsx</b>"
            ),
            init=init,
        )

    def _create_step2_card(self) -> EmptyCard:
        """创建步骤二卡片"""
        def init(widget: EmptyCard) -> None:
            btn = PushButton(FluentIcon.SAVE_AS, "注入")
            btn.clicked.connect(self._step_02_clicked)
            widget.add_widget(btn)

        return EmptyCard(
            title="第二步 - 注入修正数据",
            description=(
                "检查数据文件中的内容，确认无误后关闭文件，开始注入<br><br>"
                "<b>请注意：</b><br>"
                "• 除<b>修正列</b>以外，不要修改数据文件内的其他数据<br>"
                "• 部分格式的译文文件名中会包含类似 .zh 的语言后缀，在注入前请从文件名中移除语言后缀以正确匹配数据"
            ),
            init=init,
        )

    def _step_01_clicked(self):
        """第一步：生成修正数据"""
        try:
            # 选择输入文件夹
            if not self.input_folder:
                self.input_folder = QFileDialog.getExistingDirectory(
                    self, "选择包含结果检查文件的输入文件夹", ""
                )
                if not self.input_folder:
                    return
            
            # 选择输出文件夹
            if not self.output_folder:
                self.output_folder = QFileDialog.getExistingDirectory(
                    self, "选择输出文件夹", self.input_folder
                )
                if not self.output_folder:
                    return
            
            LogManager.get().info(f"开始生成批量修正数据：{self.input_folder}")
            
            data_dict: dict[tuple, dict] = {}
            
            # 扫描输入文件夹中的检查文件
            for entry in os.scandir(self.input_folder):
                if (entry.is_file() 
                    and self.FILE_NAME_WHITELIST.search(entry.name) is not None
                    and entry.name not in self.FILE_NAME_BLACKLIST):
                    
                    try:
                        with open(entry.path, "r", encoding="utf-8-sig") as reader:
                            json_data: dict = json.load(reader)
                            
                            if not isinstance(json_data, dict):
                                continue
                            
                            for file_path, items_by_path in json_data.items():
                                if not isinstance(items_by_path, dict):
                                    continue
                                
                                # 处理文件路径
                                chunks = file_path.split("|")
                                if len(chunks) == 1:
                                    group = self.FILE_NAME_WHITELIST.sub(r"\2", entry.name)
                                else:
                                    group = self.FILE_NAME_WHITELIST.sub(r"\2", entry.name) + " " + chunks[1].strip()
                                    file_path = chunks[0].strip()
                                
                                # 添加数据
                                for src, dst in items_by_path.items():
                                    key = (file_path, src)
                                    data_dict.setdefault(key, {})["src"] = src
                                    data_dict.setdefault(key, {})["dst"] = dst
                                    data_dict.setdefault(key, {})["group"] = data_dict.get(key, {}).get("group", []) + [group]
                                    data_dict.setdefault(key, {})["file_path"] = file_path
                    except Exception as e:
                        LogManager.get().warning(f"读取文件失败 {entry.name}: {e}")
                        continue
            
            # 有效性检查
            if len(data_dict) == 0:
                InfoBar.warning("提示", "未找到需要修正的数据，请检查输入文件夹中是否有结果检查文件", parent=self)
                return
            
            # 排序
            items = sorted(data_dict.values(), key=lambda x: (x.get("file_path", ""), str(x.get("group", ""))))
            
            # 创建 Excel 工作簿
            book = openpyxl.Workbook()
            sheet = book.active
            
            # 设置表头
            headers = ["文件名", "错误类型", "原文（勿修改此列）", "译文（勿修改此列）", "修正（请修改此列）"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
            
            # 设置列宽和筛选
            sheet.auto_filter.ref = "A1:E1"
            sheet.column_dimensions["A"].width = 12
            sheet.column_dimensions["B"].width = 12
            sheet.column_dimensions["C"].width = 64
            sheet.column_dimensions["D"].width = 64
            sheet.column_dimensions["E"].width = 64
            
            # 添加数据
            for i, item in enumerate(items, 2):
                sheet.cell(row=i, column=1, value=item.get("file_path"))
                sheet.cell(row=i, column=2, value="\n".join(item.get("group", [])))
                sheet.cell(row=i, column=3, value=item.get("src"))
                sheet.cell(row=i, column=4, value=item.get("dst"))
                sheet.cell(row=i, column=5, value=item.get("dst"))  # 默认值为译文
            
            # 保存文件
            output_path = Path(self.output_folder) / "批量修正.xlsx"
            os.makedirs(output_path.parent, exist_ok=True)
            book.save(output_path)
            
            LogManager.get().info(f"修正数据已生成: {output_path}")
            InfoBar.success(
                "任务完成", 
                f"已生成修正数据文件（共 {len(items)} 条数据）\n{output_path}", 
                parent=self
            )
            
        except Exception as e:
            LogManager.get().error(f"生成修正数据失败: {e}")
            InfoBar.error("错误", f"生成修正数据失败: {e}", parent=self)

    def _step_02_clicked(self):
        """第二步：注入修正数据"""
        try:
            # 确保有输出文件夹
            if not self.output_folder:
                self.output_folder = QFileDialog.getExistingDirectory(
                    self, "选择包含批量修正.xlsx的输出文件夹", ""
                )
                if not self.output_folder:
                    return
            
            LogManager.get().info("开始注入修正数据")
            
            # 读取 Excel 文件
            excel_path = Path(self.output_folder) / "批量修正.xlsx"
            if not excel_path.exists():
                InfoBar.warning("提示", "未找到批量修正.xlsx文件，请先执行步骤一", parent=self)
                return
            
            data_dict: dict[str, list[dict]] = {}
            book = openpyxl.load_workbook(excel_path)
            sheet = book.active
            
            if sheet.max_row == 0 or sheet.max_column == 0:
                InfoBar.warning("提示", "Excel文件为空", parent=self)
                return
            
            # 读取修正数据
            for row in range(2, sheet.max_row + 1):
                file_path = sheet.cell(row=row, column=1).value
                src = sheet.cell(row=row, column=3).value
                dst = sheet.cell(row=row, column=4).value
                fix = sheet.cell(row=row, column=5).value
                
                if file_path is None:
                    continue
                
                src = str(src) if src is not None else ""
                dst = str(dst) if dst is not None else ""
                fix = str(fix) if fix is not None else ""
                
                # 跳过无修改的行
                if fix == "" or fix == dst:
                    continue
                
                data_dict.setdefault(file_path, []).append({
                    "src": src,
                    "dst": dst,
                    "fix": fix,
                })
            
            if len(data_dict) == 0:
                InfoBar.warning("提示", "没有需要修正的数据（修正列与译文列相同）", parent=self)
                return
            
            # 统计信息
            total_files = len(data_dict)
            total_corrections = sum(len(v) for v in data_dict.values())
            LogManager.get().info(f"找到 {total_files} 个文件，共 {total_corrections} 处需要修正")

            if not self._ensure_translation_root():
                InfoBar.info("提示", "已取消注入操作。", parent=self)
                return

            applied_files = 0
            applied_changes = 0
            unmatched: list[tuple[str, list[dict]]] = []

            for rel_path, corrections in data_dict.items():
                rel_path = str(rel_path).strip()
                if not rel_path:
                    continue
                target_path = (Path(self.translation_root) / rel_path).resolve()
                if not target_path.exists():
                    LogManager.get().warning(f"未找到目标文件: {target_path}")
                    unmatched.append((rel_path, corrections))
                    continue

                try:
                    changed, applied, remaining = self._apply_corrections_to_file(target_path, corrections)
                except Exception as e:
                    LogManager.get().error(f"修正文件 {target_path} 失败: {e}")
                    unmatched.append((rel_path, corrections))
                    continue

                if applied > 0:
                    applied_changes += applied
                    applied_files += 1
                if remaining:
                    unmatched.append((rel_path, remaining))

            if applied_changes > 0:
                InfoBar.success(
                    "注入完成",
                    f"已更新 {applied_files} 个文件，共应用 {applied_changes} 处修正",
                    parent=self
                )
            else:
                InfoBar.info("提示", "未在目标文件中找到可匹配的修正项，请检查修正数据。", parent=self)

            if unmatched:
                details = "\n".join(f"{path}（未匹配 {len(items)} 项）" for path, items in unmatched[:5])
                LogManager.get().warning(f"有 {len(unmatched)} 个文件的修正未应用。\n{details}")
                InfoBar.warning(
                    "部分修正未应用",
                    f"有 {len(unmatched)} 个文件未成功注入，详情请查看日志。",
                    parent=self
                )
            
        except Exception as e:
            LogManager.get().error(f"注入修正数据失败: {e}")
            InfoBar.error("错误", f"注入修正数据失败: {e}", parent=self)

    def _auto_convert_line_break(self, src: str, fix: str) -> str:
        """根据原文换行符对修正文本中的换行符进行转换"""
        if "_x000D_" not in src and "\r" not in src:
            return fix
        else:
            return fix.replace("\n", "_x000D_\n").replace("_x000D_\n", "\r\n")

    # ===== 注入辅助方法 =====

    def _ensure_translation_root(self) -> bool:
        """确保已选择翻译文件所在目录"""
        if self.translation_root and Path(self.translation_root).exists():
            return True

        default_root = ""
        try:
            config = Config().load()
            default_root = config.renpy_tl_folder or config.renpy_game_folder or ""
        except Exception:
            pass

        folder = QFileDialog.getExistingDirectory(
            self,
            "选择翻译文件所在目录（通常为 tl/<语言> 目录）",
            default_root
        )
        if folder:
            self.translation_root = folder
            LogManager.get().info(f"已选择译文目录: {folder}")
            return True
        return False

    def _apply_corrections_to_file(self, file_path: Path, corrections: list[dict]) -> tuple[bool, int, list[dict]]:
        """根据文件类型应用修正"""
        suffix = file_path.suffix.lower()
        if suffix == ".rpy":
            return self._apply_rpy_corrections(file_path, corrections)
        else:
            return self._apply_text_corrections(file_path, corrections)

    def _apply_rpy_corrections(self, file_path: Path, corrections: list[dict]) -> tuple[bool, int, list[dict]]:
        """在 Ren'Py 翻译文件中应用修正"""
        try:
            content = file_path.read_text(encoding="utf-8")
        except Exception as e:
            LogManager.get().error(f"读取文件失败 {file_path}: {e}")
            return False, 0, corrections

        newline = "\r\n" if "\r\n" in content else "\n"
        lines = content.splitlines()

        state = [
            {
                "src": item.get("src", ""),
                "dst": item.get("dst", ""),
                "fix": item.get("fix", ""),
                "applied": False,
            }
            for item in corrections
        ]

        changed = False
        applied = 0

        i = 0
        while i < len(lines):
            stripped = lines[i].strip()
            if stripped.startswith("old "):
                original_literal = stripped[4:].strip()
                original_value = self._decode_renpy_literal(original_literal)
                j = i + 1
                while j < len(lines):
                    stripped_new = lines[j].strip()
                    if stripped_new == "":
                        j += 1
                        continue
                    if not stripped_new.startswith("new "):
                        break

                    translation_literal = stripped_new[4:].strip()
                    translation_value = self._decode_renpy_literal(translation_literal)

                    for item in state:
                        if item["applied"]:
                            continue
                        if (
                            self._normalize_text(original_value) == self._normalize_text(item["src"])
                            and self._normalize_text(translation_value) == self._normalize_text(item["dst"])
                        ):
                            fixed_text = self._auto_convert_line_break(item["src"], item["fix"])
                            encoded = self._encode_renpy_literal(fixed_text)
                            indent = lines[j][: len(lines[j]) - len(lines[j].lstrip())]
                            lines[j] = f'{indent}new "{encoded}"'
                            item["applied"] = True
                            changed = True
                            applied += 1
                            break
                    break
            i += 1

        remaining = [
            {"src": item["src"], "dst": item["dst"], "fix": item["fix"]}
            for item in state if not item["applied"]
        ]

        if changed:
            backup_path = file_path.with_suffix(file_path.suffix + ".bak")
            if not backup_path.exists():
                shutil.copy2(file_path, backup_path)
            file_path.write_text(newline.join(lines) + newline, encoding="utf-8", newline=newline)

        return changed, applied, remaining

    def _apply_text_corrections(self, file_path: Path, corrections: list[dict]) -> tuple[bool, int, list[dict]]:
        """对常规文本文件应用修正"""
        try:
            content = file_path.read_text(encoding="utf-8")
        except Exception as e:
            LogManager.get().error(f"读取文件失败 {file_path}: {e}")
            return False, 0, corrections

        changed = False
        applied = 0
        remaining: list[dict] = []

        for item in corrections:
            dst = str(item.get("dst", "") or "")
            fix = str(item.get("fix", "") or "")
            src = str(item.get("src", "") or "")
            fix_value = self._auto_convert_line_break(src, fix)

            if dst and dst in content:
                if fix_value != dst:
                    content = content.replace(dst, fix_value, 1)
                    changed = True
                applied += 1
            else:
                remaining.append(item)

        if changed:
            backup_path = file_path.with_suffix(file_path.suffix + ".bak")
            if not backup_path.exists():
                shutil.copy2(file_path, backup_path)
            file_path.write_text(content, encoding="utf-8")

        return changed, applied, remaining

    def _decode_renpy_literal(self, literal: str) -> str:
        """解析 Ren'Py old/new 行的字符串字面量"""
        text = (literal or "").strip()
        if text.startswith("u\"") or text.startswith("u'"):
            text = text[1:]
        try:
            return ast.literal_eval(text)
        except Exception:
            return text.strip('"').strip("'")

    def _encode_renpy_literal(self, text: str) -> str:
        """将文本编码为 Ren'Py 字面量"""
        if text is None:
            text = ""
        encoded = text.replace("\\", "\\\\").replace('"', '\\"')
        encoded = encoded.replace("\r\n", "\\n").replace("\r", "\\n").replace("\n", "\\n")
        return encoded

    def _normalize_text(self, text: str) -> str:
        """标准化文本以便匹配"""
        if text is None:
            return ""
        return text.replace("\r\n", "\n").replace("\r", "\n").replace("_x000D_", "")
