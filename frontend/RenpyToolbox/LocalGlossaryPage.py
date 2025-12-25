"""
本地词库管理页面
提供 Excel 导入 / 导出 与配置同步能力，方便在翻译流程中使用术语替换
"""

import dataclasses
import time
import re
from pathlib import Path
from typing import List, Dict

from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
)
from PyQt5.QtCore import QThread, pyqtSignal
from qfluentwidgets import (
    CardWidget,
    PrimaryPushButton,
    PushButton,
    InfoBar,
    FluentIcon,
    TitleLabel,
    CaptionLabel,
    StrongBodyLabel,
    isDarkTheme,
    qconfig,
)

from base.Base import Base
from module.Config import Config
from base.LogManager import LogManager
from module.Text.SkipRules import should_skip_text

try:
    from openpyxl import load_workbook, Workbook
except ImportError:  # pragma: no cover - openpyxl 在 requirements 中已声明
    load_workbook = None
    Workbook = None


class GlossaryTranslateWorker(QThread):
    progress = pyqtSignal(str, int)  # message, percent
    finished = pyqtSignal(bool, str, object)  # success, message, results(list[(row, dst)])

    def __init__(
        self,
        tasks: List[tuple[int, str]],
        *,
        source_lang: str,
        target_lang: str,
        engine: str = "google",
        parent=None,
    ):
        super().__init__(parent)
        self.tasks = tasks
        self.source_lang = source_lang
        self.target_lang = target_lang
        self.engine = engine
        self._logger = LogManager.get()

    def run(self):
        try:
            from module.Engine.FastTranslator import FastTranslator

            if not self.tasks:
                self.finished.emit(True, "No tasks", [])
                return

            self.progress.emit("正在翻译术语库...", 0)
            translator = FastTranslator(engine=self.engine)
            srcs = [src for _, src in self.tasks]
            translated = translator.translate_batch(srcs, target_lang=self.target_lang, source_lang=self.source_lang)

            results: List[tuple[int, str]] = []
            for idx, (row, _) in enumerate(self.tasks):
                dst = translated[idx] if idx < len(translated) else ""
                results.append((row, dst))

            self.progress.emit("术语库翻译完成", 100)
            self.finished.emit(True, f"Translated {len(results)} items", results)

        except Exception as exc:
            self._logger.error(f"术语库翻译失败: {exc}")
            self.finished.emit(False, str(exc), [])


class GlossaryLLMTranslateWorker(QThread):
    progress = pyqtSignal(str, int)  # message, percent
    finished = pyqtSignal(bool, str, object)  # success, message, results(list[(row, dst)])

    def __init__(
        self,
        tasks: List[tuple[int, str]],
        *,
        config: Config,
        platform: dict,
        batch_size: int = 30,
        parent=None,
    ):
        super().__init__(parent)
        self.tasks = tasks
        self.config = config
        self.platform = platform or {}
        self.batch_size = max(1, int(batch_size))
        self._logger = LogManager.get()

    @staticmethod
    def _decode_jsonline(response_text: str, expected: int) -> List[str]:
        try:
            import json_repair as repair
        except Exception:
            repair = None

        mapping: Dict[str, str] = {}
        for raw in (response_text or "").splitlines():
            line = raw.strip()
            if not line or line.startswith("```"):
                continue
            try:
                data = repair.loads(line) if repair else None  # type: ignore[attr-defined]
            except Exception:
                data = None
            if isinstance(data, dict) and len(data) == 1:
                k, v = next(iter(data.items()))
                if isinstance(k, str) and isinstance(v, str):
                    mapping[k] = v

        if not mapping:
            try:
                data = repair.loads(response_text) if repair else None  # type: ignore[attr-defined]
            except Exception:
                data = None
            if isinstance(data, dict):
                for k, v in data.items():
                    if isinstance(k, str) and isinstance(v, str):
                        mapping[k] = v

        return [mapping.get(str(i), "") for i in range(expected)]

    @staticmethod
    def _convert_chinese_form(config: Config, text: str) -> str:
        try:
            from base.BaseLanguage import BaseLanguage
        except Exception:
            return text

        if str(getattr(config, "target_language", "")).upper() != str(BaseLanguage.Enum.ZH):
            return text

        try:
            import opencc
        except Exception:
            return text

        try:
            if bool(getattr(config, "traditional_chinese_enable", False)):
                return opencc.OpenCC("s2tw").convert(text)
            return opencc.OpenCC("t2s").convert(text)
        except Exception:
            return text

    def run(self):
        try:
            if not self.tasks:
                self.finished.emit(True, "No tasks", [])
                return

            if not self.platform:
                self.finished.emit(False, "未选择翻译引擎，请先在“翻译引擎”里设置并启用一个平台。", [])
                return

            from module.Engine.TaskRequester import TaskRequester
            from module.PromptBuilder import PromptBuilder

            # 术语库翻译不注入术语库自身，避免空 dst 干扰模型
            config_for_prompt = dataclasses.replace(
                self.config,
                glossary_enable=False,
                auto_glossary_enable=False,
            )
            prompt_builder = PromptBuilder(config_for_prompt)

            all_results: List[tuple[int, str]] = []
            total = len(self.tasks)
            total_batches = (total + self.batch_size - 1) // self.batch_size

            for batch_index in range(total_batches):
                start = batch_index * self.batch_size
                batch = self.tasks[start:start + self.batch_size]
                srcs = [src for _, src in batch]

                self.progress.emit(
                    f"正在使用 LLM 翻译术语库… ({min(start, total)}/{total})",
                    int(batch_index / max(1, total_batches) * 100),
                )

                if self.platform.get("api_format") != Base.APIFormat.SAKURALLM:
                    messages, _ = prompt_builder.generate_prompt(srcs, [], [], False)
                else:
                    messages, _ = prompt_builder.generate_prompt_sakura(srcs)

                requester = TaskRequester(config_for_prompt, self.platform, batch_index)
                skip, _, response_text, _, _ = requester.request(messages)

                if skip or not response_text:
                    translated = srcs
                else:
                    translated = self._decode_jsonline(response_text, len(srcs))
                    translated = [
                        t if isinstance(t, str) and t.strip() else src
                        for t, src in zip(translated, srcs)
                    ]

                translated = [self._convert_chinese_form(config_for_prompt, t) for t in translated]

                for (row, _), dst in zip(batch, translated):
                    all_results.append((row, dst))

            self.progress.emit("术语库翻译完成", 100)
            self.finished.emit(True, f"Translated {len(all_results)} items", all_results)

        except Exception as exc:
            self._logger.error(f"术语库 LLM 翻译失败: {exc}")
            self.finished.emit(False, str(exc), [])


class LocalGlossaryPage(Base, QWidget):
    """本地词库管理页面"""

    HEADERS = ("原文", "译文", "类别", "备注")  # 类别示例：角色/地名/物品/术语
    # 过滤器关键字（参考 AiNiee NER 过滤规则），命中则跳过
    FILTER_KEYWORDS = (
        '-', '…', '一', '―', '？', '©', '章　', 'ー', 'http', '！', '=', '"', '＋', '：', '『', 'ぃ', '～',
        '♦', '〇', '└', "'", "/", "｢", "）", "（", "♥", "●", "!", "】", "【", "<", ">", "*", "〜", "EV",
        "♪", "^", "★", "※", ".", "|", "ｰ", "%", "if", "Lv", "(", "\\", "]", "[", "◆", ":", "_", "ｗｗｗ",
        "、", "ぁぁ", "んえ", "んんん",
    )

    def __init__(self, object_name: str, parent=None):
        Base.__init__(self)
        QWidget.__init__(self, parent)
        self.setObjectName(object_name)
        self.setProperty("toolboxPage", True)

        self.config = Config().load()
        self.logger = LogManager.get()
        self._translate_worker: QThread | None = None
        self._translate_llm_button: PushButton | None = None
        self._translate_fast_button: PushButton | None = None

        self._init_ui()
        self._load_from_config()

        # 监听主题变化以更新表格配色
        qconfig.themeChanged.connect(self._on_theme_changed)

    # --- UI ---
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = TitleLabel("📚 本地词库管理")
        layout.addWidget(title)

        desc = CaptionLabel(
            "支持从 Excel 导入术语表，编辑后保存到配置文件，并可导出为 Excel 共享给团队。"
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)

        layout.addWidget(self._build_toolbar_card())
        layout.addWidget(self._build_table_card())
        layout.addStretch(1)

    def _build_toolbar_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 12)
        v_layout.setSpacing(6)

        # 第一排：导入/导出/保存/加载/去重
        row1 = QHBoxLayout()
        row1.setSpacing(12)

        import_btn = PrimaryPushButton("导入 Excel", icon=FluentIcon.DOWNLOAD)
        import_btn.clicked.connect(self._on_import_excel)
        row1.addWidget(import_btn)

        export_btn = PushButton("导出 Excel", icon=FluentIcon.SHARE)
        export_btn.clicked.connect(self._on_export_excel)
        row1.addWidget(export_btn)

        save_btn = PrimaryPushButton("保存到配置", icon=FluentIcon.SAVE)
        save_btn.clicked.connect(self._save_to_config)
        row1.addWidget(save_btn)

        load_btn = PushButton("从配置加载", icon=FluentIcon.HISTORY)
        load_btn.clicked.connect(self._load_from_config)
        row1.addWidget(load_btn)

        dedup_btn = PushButton("去重重复", icon=FluentIcon.FILTER)
        dedup_btn.setToolTip("按原文去重，优先保留已有译文/类别/备注")
        dedup_btn.clicked.connect(self._deduplicate_rows)
        row1.addWidget(dedup_btn)

        row1.addStretch(1)
        v_layout.addLayout(row1)

        # 第二排：新增/删除/清空/自动分类/重新扫描角色名
        row2 = QHBoxLayout()
        row2.setSpacing(12)

        add_btn = PushButton("新增条目", icon=FluentIcon.ADD)
        add_btn.clicked.connect(self._add_row)
        row2.addWidget(add_btn)

        delete_btn = PushButton("删除选中", icon=FluentIcon.DELETE)
        delete_btn.clicked.connect(self._remove_selected_rows)
        row2.addWidget(delete_btn)

        clear_btn = PushButton("清空全部", icon=FluentIcon.CLOSE)
        clear_btn.setToolTip("删除所有术语并写入配置")
        clear_btn.clicked.connect(self._clear_all)
        row2.addWidget(clear_btn)

        auto_type_btn = PushButton("自动分类", icon=FluentIcon.TAG)
        auto_type_btn.setToolTip("先尝试 NER（需模型），再用关键词规则填充空白类别")
        auto_type_btn.clicked.connect(self._auto_classify_entries)
        row2.addWidget(auto_type_btn)

        scan_btn = PushButton("重新扫描角色名", icon=FluentIcon.SYNC)
        scan_btn.setToolTip("扫描游戏目录，自动提取角色名到术语表（清空旧的自动提取数据）")
        scan_btn.clicked.connect(self._on_rescan_characters)
        row2.addWidget(scan_btn)

        translate_llm_btn = PrimaryPushButton("LLM 翻译空译文", icon=FluentIcon.SEND)
        translate_llm_btn.setToolTip("使用已配置的翻译引擎（LLM/API）批量翻译空译文/占位译文，不会覆盖已有译文")
        translate_llm_btn.clicked.connect(self._on_translate_glossary_llm)
        row2.addWidget(translate_llm_btn)
        self._translate_llm_button = translate_llm_btn

        translate_fast_btn = PushButton("极速翻译空译文", icon=FluentIcon.GLOBE)
        translate_fast_btn.setToolTip("使用 Google/Bing 进行批量翻译（更快），不覆盖已有译文")
        translate_fast_btn.clicked.connect(self._on_translate_glossary_fast)
        row2.addWidget(translate_fast_btn)
        self._translate_fast_button = translate_fast_btn

        row2.addStretch(1)
        v_layout.addLayout(row2)

        return card

    def _build_table_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 16)
        v_layout.setSpacing(12)

        table_label = StrongBodyLabel("术语表（可直接编辑单元格）")
        v_layout.addWidget(table_label)

        self.table = QTableWidget(0, len(self.HEADERS), self)
        self.table.setHorizontalHeaderLabels(self.HEADERS)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setMinimumSectionSize(100)
        header.setDefaultSectionSize(140)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed)
        self.table.verticalHeader().setVisible(False)
        self._apply_table_theme()
        v_layout.addWidget(self.table)

        return card

    def _apply_table_theme(self) -> None:
        """根据当前主题更新表格样式"""
        if isDarkTheme():
            stylesheet = """
                QTableWidget {
                    background-color: rgb(39, 39, 39);
                    alternate-background-color: rgb(45, 45, 45);
                    color: rgb(200, 200, 200);
                    border: 1px solid rgb(55, 55, 55);
                    border-radius: 4px;
                    gridline-color: rgb(55, 55, 55);
                }
                QTableWidget::item {
                    padding: 6px;
                }
                QTableWidget::item:selected {
                    background-color: rgb(70, 70, 70);
                    color: rgb(255, 255, 255);
                }
                QHeaderView::section {
                    background-color: rgb(50, 50, 50);
                    color: rgb(200, 200, 200);
                    padding: 8px;
                    border: none;
                    border-bottom: 1px solid rgb(65, 65, 65);
                    font-weight: bold;
                }
            """
        else:
            stylesheet = """
                QTableWidget {
                    background-color: rgb(255, 255, 255);
                    alternate-background-color: rgb(248, 248, 248);
                    color: rgb(32, 32, 32);
                    border: 1px solid rgb(220, 220, 220);
                    border-radius: 4px;
                    gridline-color: rgb(230, 230, 230);
                }
                QTableWidget::item {
                    padding: 6px;
                }
                QTableWidget::item:selected {
                    background-color: rgb(210, 210, 210);
                    color: rgb(0, 0, 0);
                }
                QHeaderView::section {
                    background-color: rgb(245, 245, 245);
                    color: rgb(32, 32, 32);
                    padding: 8px;
                    border: none;
                    border-bottom: 1px solid rgb(220, 220, 220);
                    font-weight: bold;
                }
            """
        self.table.setStyleSheet(stylesheet)

    def _on_theme_changed(self) -> None:
        """主题切换时同步更新表格样式"""
        self._apply_table_theme()

    # --- 数据操作 ---
    def _add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        for col in range(len(self.HEADERS)):
            self.table.setItem(row, col, QTableWidgetItem(""))
        self.table.setCurrentCell(row, 0)

    def _remove_selected_rows(self):
        row = self.table.currentRow()
        if row < 0:
            InfoBar.warning("提示", "请选择需要删除的条目", parent=self)
            return
        self.table.removeRow(row)

    def _deduplicate_rows(self):
        """按原文去重，尽量保留已有译文/类别/备注"""
        entries = self._collect_table_data()
        if not entries:
            InfoBar.info("提示", "表格为空，暂无可去重的数据", parent=self)
            return

        key_index: Dict[str, int] = {}
        deduped: List[Dict[str, str]] = []
        for item in entries:
            key = self._normalize_src(item.get("src", ""))
            if not key:
                continue
            if key not in key_index:
                deduped.append({
                    "src": item.get("src", "").strip(),
                    "dst": item.get("dst", "").strip(),
                    "type": item.get("type", "").strip(),
                    "comment": item.get("comment", "").strip(),
                })
                key_index[key] = len(deduped) - 1
                continue

            merged = self._merge_entries(deduped[key_index[key]], item)
            deduped[key_index[key]] = merged

        removed = len(entries) - len(deduped)
        if removed > 0:
            self._set_table_data(deduped)
            InfoBar.success("完成", f"已去除重复 {removed} 条，保留 {len(deduped)} 条", parent=self)
        else:
            InfoBar.info("提示", "未发现重复条目", parent=self)

    def _clear_all(self):
        """清空表格并写回配置"""
        self.table.setRowCount(0)
        self.config = Config().load()
        self.config.glossary_data = []
        self.config.glossary_enable = False
        self.config.save()
        InfoBar.success("已清空", "已删除所有术语并写入配置", parent=self)

    @staticmethod
    def _map_language_to_fasttranslator_code(
        lang: str,
        *,
        is_target: bool,
        traditional_chinese_enable: bool,
    ) -> str:
        key = (str(lang or "").strip() or "auto")
        upper = key.upper()

        if upper in {"AUTO", "自动", "NONE"}:
            return "auto"

        if upper in {"ZH", "CHINESE", "CN", "ZH-CN", "ZH_CN", "ZH-HANS", "ZH_HANS"}:
            return "zh-TW" if is_target and traditional_chinese_enable else "zh-CN"
        if upper in {"EN", "ENGLISH"}:
            return "en"
        if upper in {"JA", "JP", "JAPANESE"}:
            return "ja"
        if upper in {"KO", "KR", "KOREAN"}:
            return "ko"
        if upper in {"RU", "RUSSIAN"}:
            return "ru"
        if upper in {"AR", "ARABIC"}:
            return "ar"
        if upper in {"DE", "GERMAN"}:
            return "de"
        if upper in {"FR", "FRENCH"}:
            return "fr"
        if upper in {"PL", "POLISH"}:
            return "pl"
        if upper in {"ES", "SPANISH"}:
            return "es"
        if upper in {"IT", "ITALIAN"}:
            return "it"
        if upper in {"PT", "PORTUGUESE"}:
            return "pt"
        if upper in {"HU", "HUNGARIAN"}:
            return "hu"
        if upper in {"TR", "TURKISH"}:
            return "tr"
        if upper in {"TH", "THAI"}:
            return "th"
        if upper in {"ID", "INDONESIAN"}:
            return "id"
        if upper in {"VI", "VIETNAMESE"}:
            return "vi"

        return key

    def _collect_glossary_translate_tasks(self) -> List[tuple[int, str]]:
        tasks: List[tuple[int, str]] = []
        rows = self.table.rowCount()
        for row in range(rows):
            src_item = self.table.item(row, 0)
            if not src_item:
                continue
            src = (src_item.text() or "").strip()
            if not src:
                continue

            dst_item = self.table.item(row, 1)
            dst = (dst_item.text() if dst_item else "").strip()
            if dst and dst != src:
                continue

            tasks.append((row, src))
        return tasks

    def _set_translate_buttons_enabled(self, enabled: bool) -> None:
        for btn in (self._translate_llm_button, self._translate_fast_button):
            if btn is not None:
                btn.setEnabled(enabled)

    def _on_translate_glossary_llm(self):
        """批量翻译术语库（LLM/API）：仅填充译文为空/等于原文的行，不覆盖已有译文。"""
        if self._translate_worker and self._translate_worker.isRunning():
            InfoBar.info("提示", "术语库翻译正在进行中，请稍候…", parent=self)
            return

        tasks = self._collect_glossary_translate_tasks()
        if not tasks:
            InfoBar.info("提示", "没有需要翻译的条目（译文列已填充）", parent=self)
            return

        config = Config().load()
        try:
            platform = config.get_platform(getattr(config, "activate_platform", 0))
        except Exception:
            platform = None
        if not platform:
            InfoBar.error("错误", "未找到可用的翻译引擎，请先在“翻译引擎”里配置并启用一个平台。", parent=self)
            return

        self._set_translate_buttons_enabled(False)

        worker = GlossaryLLMTranslateWorker(
            tasks,
            config=config,
            platform=platform,
            batch_size=30,
            parent=self,
        )
        worker.progress.connect(self._on_translate_glossary_progress)
        worker.finished.connect(self._on_translate_glossary_finished)
        self._translate_worker = worker

        InfoBar.info("开始翻译", f"正在使用 LLM 翻译 {len(tasks)} 条术语…", parent=self)
        worker.start()

    def _on_translate_glossary_fast(self):
        """批量翻译术语库：仅填充译文为空/等于原文的行，不覆盖已有译文。"""
        if self._translate_worker and self._translate_worker.isRunning():
            InfoBar.info("提示", "术语库翻译正在进行中，请稍候…", parent=self)
            return

        tasks = self._collect_glossary_translate_tasks()
        if not tasks:
            InfoBar.info("提示", "没有需要翻译的条目（译文列已填充）", parent=self)
            return

        config = Config().load()
        source_lang = self._map_language_to_fasttranslator_code(
            getattr(config, "source_language", "auto"),
            is_target=False,
            traditional_chinese_enable=bool(getattr(config, "traditional_chinese_enable", False)),
        )
        target_lang = self._map_language_to_fasttranslator_code(
            getattr(config, "target_language", "ZH"),
            is_target=True,
            traditional_chinese_enable=bool(getattr(config, "traditional_chinese_enable", False)),
        )

        self._set_translate_buttons_enabled(False)

        worker = GlossaryTranslateWorker(
            tasks,
            source_lang=source_lang,
            target_lang=target_lang,
            engine="alibaba",
            parent=self,
        )
        worker.progress.connect(self._on_translate_glossary_progress)
        worker.finished.connect(self._on_translate_glossary_finished)
        self._translate_worker = worker

        InfoBar.info("开始翻译", f"正在翻译 {len(tasks)} 条术语…", parent=self)
        worker.start()

    def _on_translate_glossary_progress(self, message: str, percent: int):
        self.logger.info(f"[GlossaryTranslate] {percent}% {message}")

    def _on_translate_glossary_finished(self, success: bool, message: str, results):
        self._set_translate_buttons_enabled(True)

        worker = self._translate_worker
        self._translate_worker = None
        if worker is not None:
            worker.deleteLater()

        if not success:
            InfoBar.error("翻译失败", message, parent=self)
            return

        applied = 0
        for row, dst in results or []:
            if row < 0 or row >= self.table.rowCount():
                continue
            dst_text = (dst or "").strip()
            if not dst_text:
                continue

            src_item = self.table.item(row, 0)
            src_text = (src_item.text() if src_item else "").strip()

            dst_item = self.table.item(row, 1)
            current_dst = (dst_item.text() if dst_item else "").strip()
            if current_dst and current_dst != src_text:
                continue  # 不覆盖已有译文

            if dst_item is None:
                dst_item = QTableWidgetItem("")
                self.table.setItem(row, 1, dst_item)

            dst_item.setText(dst_text)
            applied += 1

        if applied:
            InfoBar.success("翻译完成", f"已填充 {applied} 条译文（别忘了点击“保存到配置”）", parent=self)
        else:
            InfoBar.info("翻译完成", "翻译已结束，但没有产生可用译文（可能接口返回原文）", parent=self)

    def _load_from_config(self):
        data = getattr(self.config, "glossary_data", []) or []
        converted = []
        for item in data:
            if isinstance(item, dict):
                converted.append(
                    {
                        "src": item.get("src", ""),
                        "dst": item.get("dst", ""),
                        "type": item.get("type", item.get("category", "")),
                        "comment": item.get("comment", ""),
                    }
                )
        self._set_table_data(converted)
        InfoBar.success("完成", f"已从配置加载 {len(converted)} 条术语", parent=self)

    def _save_to_config(self):
        entries = self._collect_table_data()
        self.config = Config().load()
        self.config.glossary_data = entries
        self.config.glossary_enable = True if entries else self.config.glossary_enable
        self.config.save()
        InfoBar.success("保存成功", f"已写入 {len(entries)} 条术语到配置", parent=self)

    def _on_import_excel(self):
        if load_workbook is None:
            InfoBar.error("错误", "未安装 openpyxl，无法导入 Excel", parent=self)
            return

        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择术语 Excel 文件",
            "",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            workbook = load_workbook(path)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            header_map = self._build_header_map(headers)
            if "src" not in header_map or "dst" not in header_map:
                raise ValueError("未找到“原文/译文”列，请确认模板。")

            items: List[Dict[str, str]] = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                src = self._safe_cell(row, header_map.get("src"))
                dst = self._safe_cell(row, header_map.get("dst"))
                type_ = self._safe_cell(row, header_map.get("type"))
                comment = self._safe_cell(row, header_map.get("comment"))
                if not src:
                    continue
                items.append({"src": src, "dst": dst, "type": type_, "comment": comment})

            self._set_table_data(items)
            InfoBar.success("导入成功", f"已导入 {len(items)} 条术语", parent=self)
        except Exception as e:
            self.logger.error(f"导入术语失败: {e}")
            InfoBar.error("错误", f"导入失败: {e}", parent=self)

    def _on_export_excel(self):
        if Workbook is None:
            InfoBar.error("错误", "未安装 openpyxl，无法导出 Excel", parent=self)
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "保存术语 Excel",
            "",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        entries = self._collect_table_data()
        if not entries:
            InfoBar.warning("提示", "当前表格为空，未导出文件", parent=self)
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Glossary"
            sheet.append(list(self.HEADERS))
            for item in entries:
                sheet.append([item.get("src", ""), item.get("dst", ""), item.get("type", ""), item.get("comment", "")])
            workbook.save(path)
            InfoBar.success("导出成功", f"已保存到 {path}", parent=self)
        except Exception as e:
            self.logger.error(f"导出术语失败: {e}")
            InfoBar.error("错误", f"导出失败: {e}", parent=self)

    # --- 工具方法 ---
    def _set_table_data(self, items: List[Dict[str, str]]):
        self.table.setRowCount(0)
        for item in items:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(item.get("src", "")))
            self.table.setItem(row, 1, QTableWidgetItem(item.get("dst", "")))
            self.table.setItem(row, 2, QTableWidgetItem(item.get("type", "")))
            self.table.setItem(row, 3, QTableWidgetItem(item.get("comment", "")))

    def _collect_table_data(self) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        rows = self.table.rowCount()
        for row in range(rows):
            src_item = self.table.item(row, 0)
            dst_item = self.table.item(row, 1)
            type_item = self.table.item(row, 2)
            comment_item = self.table.item(row, 3)
            src = (src_item.text() if src_item else "").strip()
            dst = (dst_item.text() if dst_item else "").strip()
            type_ = (type_item.text() if type_item else "").strip()
            comment = (comment_item.text() if comment_item else "").strip()
            if not src:
                continue
            results.append({"src": src, "dst": dst, "type": type_, "comment": comment})
        return results

    @staticmethod
    def _normalize_src(text: str) -> str:
        if not text:
            return ""
        normalized = re.sub(r"\s+", " ", text)
        normalized = normalized.strip().strip("\"'“”‘’")
        return normalized.lower()

    @staticmethod
    def _merge_entries(base: Dict[str, str], incoming: Dict[str, str]) -> Dict[str, str]:
        def _clean(value: str) -> str:
            return value.strip() if isinstance(value, str) else ""

        merged = {
            "src": _clean(base.get("src")),
            "dst": _clean(base.get("dst")),
            "type": _clean(base.get("type")),
            "comment": _clean(base.get("comment")),
        }
        incoming_cleaned = {
            "src": _clean(incoming.get("src")),
            "dst": _clean(incoming.get("dst")),
            "type": _clean(incoming.get("type")),
            "comment": _clean(incoming.get("comment")),
        }

        if incoming_cleaned["dst"]:
            if not merged["dst"] or (merged["src"] and merged["dst"].lower() == merged["src"].lower()):
                merged["dst"] = incoming_cleaned["dst"]

        if incoming_cleaned["type"] and not merged["type"]:
            merged["type"] = incoming_cleaned["type"]

        if incoming_cleaned["comment"]:
            if not merged["comment"]:
                merged["comment"] = incoming_cleaned["comment"]
            elif incoming_cleaned["comment"] not in merged["comment"] and len(incoming_cleaned["comment"]) > len(merged["comment"]):
                merged["comment"] = incoming_cleaned["comment"]

        if incoming_cleaned["src"] and not merged["src"]:
            merged["src"] = incoming_cleaned["src"]

        return merged

    @staticmethod
    def _build_header_map(headers: List[str]) -> Dict[str, int]:
        alias = {
            "src": {"原文", "原始文本", "source", "src"},
            "dst": {"译文", "翻译", "target", "translation", "dst"},
            "type": {"类别", "分类", "type", "category"},
            "comment": {"备注", "说明", "comment", "note", "备注信息"},
        }
        mapping = {}
        for index, name in enumerate(headers):
            lower_name = name.lower()
            for key, options in alias.items():
                if lower_name in {opt.lower() for opt in options} and key not in mapping:
                    mapping[key] = index
        return mapping

    @staticmethod
    def _safe_cell(row, index: int) -> str:
        if index is None:
            return ""
        if index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()

    def _on_rescan_characters(self):
        """重新扫描游戏目录，提取角色名到术语表（清空旧的自动提取数据）"""
        import re
        from pathlib import Path
        from module.Text.SkipRules import should_skip_text
        
        # 重新加载配置以获取最新的游戏目录
        self.config = Config().load()
        
        # 获取游戏目录
        game_folder = self.config.renpy_game_folder
        if not game_folder:
            # 尝试让用户选择一次游戏目录
            folder = QFileDialog.getExistingDirectory(self, "选择游戏目录（包含 game 子目录）")
            if folder:
                game_folder = folder
                self.config.renpy_game_folder = game_folder
                self.config.save()
                InfoBar.info("提示", f"已设置游戏目录为: {game_folder}", parent=self)
            else:
                InfoBar.warning("警告", "请先选择游戏目录", parent=self)
                return
            
        game_path = Path(game_folder) / "game"
        if not game_path.exists():
            game_path = Path(game_folder)
            if not game_path.exists():
                InfoBar.error("错误", f"游戏目录不存在: {game_folder}", parent=self)
                return
        
        found_names = set()
        cache_key = str(game_path.resolve())
        auto_cache = dict(getattr(self.config, "glossary_auto_scan_cache", {}) or {})
        auto_cache.pop(cache_key, None)
        self.config.glossary_auto_scan_cache = auto_cache

        tl_name = getattr(self.config, "renpy_tl_folder", "") or "chinese"

        # 方法1: 从 miss_ready_replace 文件提取
        miss_names = self._extract_names_from_miss_files(game_path, tl_name)
        if miss_names:
            self.logger.info(f"Miss file contributed {len(miss_names)} candidates")
            found_names |= miss_names

        # 方法2: 直接从游戏源码扫描 Character 定义
        source_names = self._extract_names_from_source(game_path)
        if source_names:
            self.logger.info(f"Source scan contributed {len(source_names)} candidates")
            found_names |= source_names

        self.logger.info(f"Character scan found {len(found_names)} total candidates")

        # 如果仍未找到，则提示
        if not found_names:
            InfoBar.info("提示", "未找到角色名，请确认游戏目录正确", parent=self)
            return
        
        # 保留手动添加的条目，清除旧的自动提取数据
        manual_entries = []
        if self.config.glossary_data:
            for item in self.config.glossary_data:
                if isinstance(item, dict):
                    info = item.get("info", "") or item.get("comment", "")
                    if "自动提取" in info and ("character" in info.lower() or "角色" in info):
                        continue
                    manual_entries.append(item)
        
        # 添加新扫描到的
        existing_src = set(item.get("src", "") for item in manual_entries if isinstance(item, dict))
        new_entries = []
        for name in found_names:
            if name not in existing_src:
                cleaned = self._clean_text_for_classify(name)
                if not cleaned or should_skip_text(cleaned):
                    continue
                # 智能分类，默认空
                type_guess = self._categorize_term(cleaned, default="")
                new_entries.append({
                    "src": cleaned,
                    "dst": "",
                    "info": "角色名 (自动提取)",
                    "type": type_guess
                })

        # 保存到配置
        self.config.glossary_data = manual_entries + new_entries
        self.config.glossary_enable = True
        auto_cache[cache_key] = time.time()
        self.config.glossary_auto_scan_cache = auto_cache
        self.config.save()
        
        # 刷新表格
        self._load_from_config()
        InfoBar.success("完成", f"已扫描到 {len(new_entries)} 个新角色名，已清除旧的自动提取数据", parent=self)

    def _is_probable_name(self, text: str) -> bool:
        """更严格的人名判定：短、无句号、少量单词、首字母大写或全大写、允许少量连接词"""
        if not text:
            return False
        if len(text) > 40 or "\n" in text:
            return False
        # 去掉花括号标签等
        t = self._clean_text_for_classify(text)
        if not t:
            return False
        # 拒绝带句号/问号/感叹号/数字
        if any(p in t for p in (".", "?", "!", ":", "；", "。", "！", "？")):
            return False
        if any(ch.isdigit() for ch in t):
            return False
        words = t.split()
        if not words or len(words) > 4:
            return False
        allow_connectors = {"of", "the", "and"}
        cjk = any("\u3040" <= ch <= "\u30ff" or "\u4e00" <= ch <= "\u9fff" for ch in t)
        if cjk:
            # CJK 名字：不含空格或 2-3 词
            return True
        # 英文：每个词首字母大写或全大写，允许 of/the/and 小写
        for w in words:
            if w.lower() in allow_connectors:
                continue
            if not (w[:1].isupper() or w.isupper()):
                return False
        return True

    # ---- 智能分类 ----
    def _auto_categorize_entries(self, silent: bool = False) -> int:
        """为缺少类别的条目自动填充类别（关键词规则）"""
        rows = self.table.rowCount()
        changed = 0
        for row in range(rows):
            src_item = self.table.item(row, 0)
            type_item = self.table.item(row, 2)
            if not src_item:
                continue
            if type_item and type_item.text().strip():
                continue
            cleaned = self._clean_text_for_classify(src_item.text())
            if not cleaned:
                continue
            guess = self._categorize_term(cleaned)
            if guess:
                if not type_item:
                    type_item = QTableWidgetItem("")
                    self.table.setItem(row, 2, type_item)
                type_item.setText(guess)
                changed += 1
        if not silent:
            if changed:
                InfoBar.success("完成", f"已为 {changed} 条填充类别", parent=self)
            else:
                InfoBar.info("提示", "没有需要填充的类别或未找到匹配", parent=self)
        return changed

    @staticmethod
    def _categorize_term(text: str, default: str = "") -> str:
        """基于关键词/形态的简易分类"""
        if not text:
            return default
        t = text.strip()
        lower = t.lower()
        place_keywords = [
            "city", "village", "town", "forest", "mountain", "hill", "park", "garden",
            "school", "academy", "college", "campus", "church", "temple", "shrine",
            "castle", "tower", "dungeon", "cave", "ruins", "harbor", "port", "station",
            "beach", "island", "lake", "river", "bridge", "street", "road", "avenue",
            "hotel", "inn", "bar", "cafe", "shop", "market", "library"
        ]
        item_keywords = [
            "sword", "blade", "dagger", "bow", "gun", "rifle", "pistol", "armor", "shield",
            "ring", "necklace", "amulet", "bracelet", "crown", "helmet", "boots", "gloves",
            "potion", "elixir", "herb", "scroll", "book", "map", "key", "card", "ticket",
            "coin", "gem", "crystal", "stone", "orb", "staff", "wand", "medal"
        ]
        # 地名关键词匹配
        if any(k in lower for k in place_keywords):
            return "地名"
        # 物品关键词匹配
        if any(k in lower for k in item_keywords):
            return "物品"
        # 大写单词串通常为专名（角色/组织/作品）
        words = t.split()
        if words and all(w[:1].isupper() for w in words if w):
            return default or ""
        return default

    # ---- NER 分类（需本地模型） ----
    def _ner_categorize_entries(self, silent: bool = False) -> int:
        """使用本地 spaCy NER 模型为空白类别填充"""
        try:
            import spacy
        except Exception as e:
            if not silent:
                InfoBar.error("错误", f"未安装 spaCy：{e}", parent=self)
            return 0

        model_path = self._find_ner_model_path()
        if not model_path:
            if not silent:
                InfoBar.warning("提示", "未找到 NER 模型（Resource/Models/ner/*），已跳过", parent=self)
            return 0

        try:
            nlp = spacy.load(str(model_path), exclude=["parser", "tagger", "lemmatizer", "attribute_ruler", "tok2vec"])
        except Exception as e:
            if not silent:
                InfoBar.error("错误", f"加载 NER 模型失败: {e}", parent=self)
            return 0

        label_map = {
            "PER": "角色", "PERSON": "角色", "PER_NO": "角色",
            "LOC": "地名", "GPE": "地名",
            "ORG": "组织",
            "FAC": "地名",
            "PRODUCT": "物品", "ITEM": "物品",
        }

        rows = self.table.rowCount()
        changed = 0
        for row in range(rows):
            src_item = self.table.item(row, 0)
            type_item = self.table.item(row, 2)
            if not src_item:
                continue
            if type_item and type_item.text().strip():
                continue
            text = self._clean_text_for_classify(src_item.text())
            if not text:
                continue
            # 过滤包含无效关键词的条目
            if any(k in text for k in self.FILTER_KEYWORDS):
                continue
            doc = nlp(text)
            guessed = ""
            # 优先精确匹配实体文本
            for ent in doc.ents:
                if any(k in ent.text for k in self.FILTER_KEYWORDS):
                    continue
                if ent.text.strip().lower() == text.strip().lower():
                    guessed = label_map.get(ent.label_, ent.label_)
                    break
            # 否则取首个实体
            if not guessed and doc.ents:
                for ent in doc.ents:
                    if any(k in ent.text for k in self.FILTER_KEYWORDS):
                        continue
                    guessed = label_map.get(ent.label_, ent.label_)
                    if guessed:
                        break
            if guessed:
                if not type_item:
                    type_item = QTableWidgetItem("")
                    self.table.setItem(row, 2, type_item)
                type_item.setText(guessed)
                changed += 1

        if changed:
            if not silent:
                InfoBar.success("完成", f"NER 填充了 {changed} 条类别", parent=self)
        else:
            if not silent:
                InfoBar.info("提示", "未找到可填充的类别", parent=self)
        return changed

    def _find_ner_model_path(self) -> Path | None:
        """查找本地 NER 模型路径（Resource/Models/ner），按语言偏好选择"""
        candidates = []
        for base in [Path("."), Path(__file__).resolve().parents[2]]:
            model_root = (base / "Resource" / "Models" / "ner").resolve()
            if model_root.exists():
                for p in model_root.iterdir():
                    if p.is_dir() and (p / "meta.json").exists():
                        candidates.append(p)
        if not candidates:
            return None

        preferred = self._guess_ner_preference()

        def _score(path: Path) -> int:
            name = path.name.lower()
            if preferred == "ja":
                if name.startswith("ja_core"):
                    return 0
            if preferred == "en":
                if name.startswith("en_core_web_md"):
                    return 0
                if name.startswith("en_core_web_"):
                    return 1
            # 次优：其余语言/模型
            return 5

        candidates.sort(key=_score)
        return candidates[0]

    def _guess_ner_preference(self) -> str:
        """根据表格文本粗略判断偏好（ja vs en）"""
        cjk = 0
        latin = 0
        rows = min(self.table.rowCount(), 200)
        for row in range(rows):
            src_item = self.table.item(row, 0)
            if not src_item:
                continue
            text = src_item.text() or ""
            for ch in text:
                if "\u4e00" <= ch <= "\u9fff" or "\u3040" <= ch <= "\u30ff":
                    cjk += 1
                elif ch.isalpha():
                    latin += 1
        if cjk > latin:
            return "ja"
        return "en"

    def _ner_extract_names_from_game(self, game_path: Path) -> set[str]:
        """在未找到角色名时，使用 NER 从游戏源码扫描人名"""
        try:
            import spacy
        except Exception:
            return set()

        model_path = self._find_ner_model_path()
        if not model_path:
            return set()
        try:
            nlp = spacy.load(str(model_path), exclude=["parser", "tagger", "lemmatizer", "attribute_ruler", "tok2vec"])
        except Exception:
            return set()

        label_map = {"PER", "PERSON", "PER_NO"}
        names = set()
        for rpy_file in game_path.rglob("*.rpy"):
            # 跳过 tl 目录
            if "tl" in rpy_file.parts:
                continue
            try:
                for line in rpy_file.read_text(encoding="utf-8", errors="ignore").splitlines():
                    text = self._clean_text_for_classify(line)
                    if not text or len(text) > 200:
                        continue
                    if should_skip_text(text):
                        continue
                    doc = nlp(text)
                    for ent in doc.ents:
                        if ent.label_ in label_map:
                            candidate = ent.text.strip()
                            candidate = self._clean_text_for_classify(candidate)
                            if candidate and not should_skip_text(candidate) and self._is_probable_name(candidate):
                                names.add(candidate)
            except Exception:
                continue
        return names

    def _extract_names_from_source(self, game_path: Path) -> set[str]:
        """直接从游戏源码扫描 Character 定义提取角色名"""
        import re
        from module.Text.SkipRules import should_skip_text
        
        names = set()
        
        # 正则匹配 Character("Name") 或 Character(_("Name")) 或 define xxx = Character("Name")
        RE_CHARACTER_CALL = re.compile(
            r"Character\s*\(\s*(?:_\(\s*)?(['\"])((?:\\\1|.)*?)\1",
            re.MULTILINE
        )
        
        # 扫描 .rpy 文件（排除 tl 目录和 cache）
        try:
            for rpy_file in game_path.rglob("*.rpy"):
                # 跳过 tl 目录和缓存
                rel_parts = rpy_file.relative_to(game_path).parts
                if any(p in ("tl", "cache", "__pycache__") for p in rel_parts):
                    continue
                    
                try:
                    content = rpy_file.read_text(encoding="utf-8", errors="ignore")
                    for match in RE_CHARACTER_CALL.finditer(content):
                        raw_name = match.group(2)
                        # 处理转义
                        name = raw_name.replace('\\"', '"').replace("\\'", "'").replace("\\n", " ").strip()
                        
                        if not name:
                            continue
                        
                        # 跳过变量引用如 [player_name]
                        if name.startswith("[") and name.endswith("]"):
                            continue
                        
                        # 跳过过长的文本（通常不是角色名）
                        if len(name) > 50:
                            continue
                        
                        # 跳过纯数字或特殊字符
                        if name.isdigit() or all(c in "!@#$%^&*()_+-=[]{}|;:'\",.<>?/\\" for c in name):
                            continue
                        
                        # 清理并验证
                        cleaned = self._clean_text_for_classify(name)
                        if cleaned and len(cleaned) >= 2:
                            # 不使用 should_skip_text，因为角色名可能触发误判
                            if self._is_probable_name(cleaned):
                                names.add(cleaned)
                            elif len(cleaned) <= 20:  # 短文本也可能是角色名
                                names.add(cleaned)
                except Exception as e:
                    self.logger.debug(f"Error reading {rpy_file}: {e}")
                    continue
                    
        except Exception as e:
            self.logger.warning(f"Error scanning source files: {e}")
        
        self.logger.info(f"[Glossary] source scan yielded {len(names)} character names")
        return names

    def _extract_names_from_miss_files(self, game_path: Path, tl_name: str) -> set[str]:
        """从 miss_ready_replace*.txt 提取可能的人名"""
        names = set()
        tl_root = game_path / "tl" / tl_name
        if not tl_root.exists():
            self.logger.info(f"[Glossary] tl path not found: {tl_root}")
            return names
        candidates: list[Path] = []
        for base in (tl_root, tl_root / "miss"):
            candidates.extend(base.glob("miss_ready_replace*.rpy"))
            candidates.extend(base.glob("miss_ready_replace*.txt"))
        if not candidates:
            self.logger.info(f"[Glossary] no miss_ready_replace files under {tl_root}")
            return names
        for miss_file in candidates:
            try:
                self.logger.info(f"[Glossary] reading miss file: {miss_file}")
                for line in miss_file.read_text(encoding="utf-8", errors="ignore").splitlines():
                    line = line.strip()
                    if line.startswith("old "):
                        m = re.search(r'old\s+"(.*)"', line)
                        if not m:
                            continue
                        raw = m.group(1)
                        text = raw.replace('\\"', '"').replace("\\n", "\n").replace("\\\\", "\\")
                        clean = self._clean_text_for_classify(text)
                        if not clean:
                            continue
                        # 对角色名提取放宽：仅用人名判定，不再应用 should_skip_text，避免误杀
                        if self._is_probable_name(clean):
                            names.add(clean)
            except Exception:
                continue
        self.logger.info(f"[Glossary] miss files yielded {len(names)} names")
        return names

    def _auto_classify_entries(self):
        """一键分类：先 NER，再关键词兜底"""
        ner_count = self._ner_categorize_entries(silent=True)
        kw_count = self._auto_categorize_entries(silent=True)
        if ner_count or kw_count:
            InfoBar.success("完成", f"NER 填充 {ner_count} 条，关键词填充 {kw_count} 条", parent=self)
        else:
            InfoBar.info("提示", "未找到可填充的类别（可检查模型或文本内容）", parent=self)

    # ---- 文本清洗 ----
    @staticmethod
    def _clean_text_for_classify(text: str) -> str:
        """去除格式标签/空白，用于分类和过滤"""
        if not text:
            return ""
        import re
        cleaned = re.sub(r"\{/?[^}]+\}", "", text)
        cleaned = cleaned.replace("\u3000", " ").strip()
        return cleaned
