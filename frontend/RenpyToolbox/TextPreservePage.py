"""
文本保留管理页面
管理不需要翻译的文本（如专有名词、代码片段等），这些内容将在翻译过程中保持原文。
"""

from typing import List, Dict
from pathlib import Path

from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
)
from PyQt5.QtCore import Qt
from qfluentwidgets import (
    CardWidget,
    PrimaryPushButton,
    PushButton,
    InfoBar,
    FluentIcon,
    TitleLabel,
    CaptionLabel,
    StrongBodyLabel,
    isDarkTheme,
    qconfig,
)

from base.Base import Base
from module.Config import Config
from base.LogManager import LogManager
from module.Localizer.Localizer import Localizer
from frontend.RenpyToolbox.RuleStatisticsWorker import RuleStatisticsWorker

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None
    Workbook = None


class TextPreservePage(Base, QWidget):
    """文本保留管理页面"""

    HEADERS = ("原文", "备注", "命中数")
    STATS_COLUMN = 2
    STATS_COLUMN_WIDTH = 88

    @staticmethod
    def _display_headers() -> tuple[str, str, str]:
        return (
            Localizer.get().proofreading_page_col_src,
            Localizer.get().local_glossary_notes,
            Localizer.get().local_glossary_hits,
        )

    def __init__(self, object_name: str, parent=None):
        Base.__init__(self)
        QWidget.__init__(self, parent)
        self.setObjectName(object_name)
        self.setProperty("toolboxPage", True)

        self.config = Config().load()
        self.logger = LogManager.get()
        self._statistics_worker = None
        self._statistics_button: PushButton | None = None
        self._statistics_snapshot_keys: List[str] = []

        self._init_ui()
        self._load_from_config()

        # 监听主题变化以更新表格配色
        qconfig.themeChanged.connect(self._on_theme_changed)

    # --- UI ---
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = TitleLabel(Localizer.get().text_preserve_do_not_translate)
        layout.addWidget(title)

        desc = CaptionLabel(Localizer.get().text_preserve_manage_text_should_remain_unchanged_during_translation)
        desc.setWordWrap(True)
        layout.addWidget(desc)

        layout.addWidget(self._build_toolbar_card())
        layout.addWidget(self._build_table_card())
        layout.addStretch(1)

    def _build_toolbar_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 12)
        v_layout.setSpacing(6)

        # 第一排：导入/导出/保存/加载
        row1 = QHBoxLayout()
        row1.setSpacing(12)

        import_btn = PrimaryPushButton(
            Localizer.get().local_glossary_import_excel,
            icon=FluentIcon.DOWNLOAD,
        )
        import_btn.clicked.connect(self._on_import_excel)
        row1.addWidget(import_btn)

        export_btn = PushButton(
            Localizer.get().local_glossary_export_excel,
            icon=FluentIcon.SHARE,
        )
        export_btn.clicked.connect(self._on_export_excel)
        row1.addWidget(export_btn)

        save_btn = PrimaryPushButton(
            Localizer.get().text_preserve_save_settings,
            icon=FluentIcon.SAVE,
        )
        save_btn.clicked.connect(self._save_to_config)
        row1.addWidget(save_btn)

        load_btn = PushButton(
            Localizer.get().text_preserve_load_settings,
            icon=FluentIcon.HISTORY,
        )
        load_btn.clicked.connect(self._load_from_config)
        row1.addWidget(load_btn)

        row1.addStretch(1)
        v_layout.addLayout(row1)

        # 第二排：新增/删除/清空/重新扫描
        row2 = QHBoxLayout()
        row2.setSpacing(12)

        add_btn = PushButton(Localizer.get().local_glossary_add_entry, icon=FluentIcon.ADD)
        add_btn.clicked.connect(self._add_row)
        row2.addWidget(add_btn)

        delete_btn = PushButton(
            Localizer.get().local_glossary_delete_selected,
            icon=FluentIcon.DELETE,
        )
        delete_btn.clicked.connect(self._remove_selected_rows)
        row2.addWidget(delete_btn)

        dedup_btn = PushButton(Localizer.get().text_preserve_deduplicate, icon=FluentIcon.FILTER)
        dedup_btn.setToolTip(Localizer.get().text_preserve_deduplicate_source_text_merge_notes_prefer_rows)
        dedup_btn.clicked.connect(self._deduplicate_rows)
        row2.addWidget(dedup_btn)

        clear_btn = PushButton(Localizer.get().local_glossary_clear_all, icon=FluentIcon.CLOSE)
        clear_btn.setToolTip(Localizer.get().text_preserve_delete_all_do_not_translate_entries_save)
        clear_btn.clicked.connect(self._clear_all)
        row2.addWidget(clear_btn)

        statistics_btn = PushButton(
            Localizer.get().local_glossary_count_hits,
            icon=FluentIcon.SEARCH,
        )
        statistics_btn.setToolTip(Localizer.get().text_preserve_count_how_many_cached_output_entries_match)
        statistics_btn.clicked.connect(self._on_statistics_clicked)
        row2.addWidget(statistics_btn)
        self._statistics_button = statistics_btn

        scan_btn = PushButton(
            Localizer.get().text_preserve_rescan_variables,
            icon=FluentIcon.SYNC,
        )
        scan_btn.setToolTip(Localizer.get().text_preserve_scan_game_folder_variable_references_replace_previous)
        scan_btn.clicked.connect(self._on_rescan_variables)
        row2.addWidget(scan_btn)

        row2.addStretch(1)
        v_layout.addLayout(row2)

        return card

    def _build_table_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 16)
        v_layout.setSpacing(12)

        table_label = StrongBodyLabel(Localizer.get().text_preserve_do_not_translate_entries_cells_editable)
        v_layout.addWidget(table_label)

        self.table = QTableWidget(0, len(self.HEADERS), self)
        self.table.setHorizontalHeaderLabels(self._display_headers())
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(self.STATS_COLUMN, QHeaderView.Fixed)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed)
        self.table.verticalHeader().setVisible(False)
        self.table.setColumnWidth(self.STATS_COLUMN, self.STATS_COLUMN_WIDTH)
        self.table.itemChanged.connect(self._on_table_item_changed)
        self._apply_table_theme()
        v_layout.addWidget(self.table)

        return card

    def _apply_table_theme(self) -> None:
        """根据当前主题更新表格样式"""
        if isDarkTheme():
            stylesheet = """
                QTableWidget {
                    background-color: rgb(39, 39, 39);
                    alternate-background-color: rgb(45, 45, 45);
                    color: rgb(200, 200, 200);
                    border: 1px solid rgb(55, 55, 55);
                    border-radius: 4px;
                    gridline-color: rgb(55, 55, 55);
                }
                QTableWidget::item {
                    padding: 6px;
                }
                QTableWidget::item:selected {
                    background-color: rgb(70, 70, 70);
                    color: rgb(255, 255, 255);
                }
                QHeaderView::section {
                    background-color: rgb(50, 50, 50);
                    color: rgb(200, 200, 200);
                    padding: 8px;
                    border: none;
                    border-bottom: 1px solid rgb(65, 65, 65);
                    font-weight: bold;
                }
            """
        else:
            stylesheet = """
                QTableWidget {
                    background-color: rgb(255, 255, 255);
                    alternate-background-color: rgb(248, 248, 248);
                    color: rgb(32, 32, 32);
                    border: 1px solid rgb(220, 220, 220);
                    border-radius: 4px;
                    gridline-color: rgb(230, 230, 230);
                }
                QTableWidget::item {
                    padding: 6px;
                }
                QTableWidget::item:selected {
                    background-color: rgb(210, 210, 210);
                    color: rgb(0, 0, 0);
                }
                QHeaderView::section {
                    background-color: rgb(245, 245, 245);
                    color: rgb(32, 32, 32);
                    padding: 8px;
                    border: none;
                    border-bottom: 1px solid rgb(220, 220, 220);
                    font-weight: bold;
                }
            """
        self.table.setStyleSheet(stylesheet)

    def _on_theme_changed(self) -> None:
        """主题切换时同步更新表格样式"""
        self._apply_table_theme()

    def _create_table_item(self, text: str = "", *, editable: bool = True) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setTextAlignment(Qt.AlignCenter)
        return item

    def _build_statistics_entry_key(self, item: Dict[str, str]) -> str:
        return self._normalize_src(str(item.get("src", "") or ""))

    def _set_statistics_buttons_enabled(self, enabled: bool) -> None:
        if self._statistics_button is not None:
            self._statistics_button.setEnabled(enabled)

    def _invalidate_statistics(self) -> None:
        self._statistics_snapshot_keys = []
        self.table.blockSignals(True)
        try:
            for row in range(self.table.rowCount()):
                item = self.table.item(row, self.STATS_COLUMN)
                if item is None:
                    item = self._create_table_item("", editable=False)
                    self.table.setItem(row, self.STATS_COLUMN, item)
                else:
                    item.setText("")
        finally:
            self.table.blockSignals(False)

    def _on_table_item_changed(self, item: QTableWidgetItem) -> None:
        if item is None:
            return
        if item.column() != 0:
            return
        self._invalidate_statistics()

    # --- 数据操作 ---
    def _add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, self._create_table_item(""))
        self.table.setItem(row, 1, self._create_table_item(""))
        self.table.setItem(row, self.STATS_COLUMN, self._create_table_item("", editable=False))
        self.table.setCurrentCell(row, 0)
        self._invalidate_statistics()

    def _remove_selected_rows(self):
        row = self.table.currentRow()
        if row < 0:
            InfoBar.warning(
                Localizer.get().notice,
                Localizer.get().local_glossary_select_entry_delete,
                parent=self,
            )
            return
        self.table.removeRow(row)
        self._invalidate_statistics()

    def _deduplicate_rows(self):
        """按原文去重，优先保留有备注的条目"""
        entries = self._collect_table_data()
        if not entries:
            InfoBar.info(
                Localizer.get().notice,
                Localizer.get().local_glossary_table_empty,
                parent=self,
            )
            return

        key_index: Dict[str, int] = {}
        deduped: List[Dict[str, str]] = []
        for item in entries:
            key = self._normalize_src(item.get("src", ""))
            if not key:
                continue
            if key not in key_index:
                deduped.append({"src": item.get("src", "").strip(), "comment": item.get("comment", "").strip()})
                key_index[key] = len(deduped) - 1
            else:
                existing = deduped[key_index[key]]
                merged = self._merge_entries(existing, item)
                deduped[key_index[key]] = merged

        removed = len(entries) - len(deduped)
        if removed > 0:
            self._set_table_data(deduped)
            InfoBar.success(
                Localizer.get().local_glossary_completed,
                Localizer.get().local_glossary_removed_duplicate_entries_kept.format(removed=removed, deduped_count=len(deduped)),
                parent=self,
            )
        else:
            InfoBar.info(
                Localizer.get().notice,
                Localizer.get().local_glossary_no_duplicate_entries_found,
                parent=self,
            )

    def _clear_all(self):
        """清空表格并写回配置"""
        self.table.setRowCount(0)
        self.config = Config().load()
        self.config.text_preserve_data = []
        self.config.text_preserve_enable = False
        self.config.save()
        self._invalidate_statistics()
        InfoBar.success(
            Localizer.get().local_glossary_cleared,
            Localizer.get().text_preserve_deleted_all_do_not_translate_entries_saved,
            parent=self,
        )

    def _load_from_config(self):
        data = getattr(self.config, "text_preserve_data", []) or []
        converted = []
        for item in data:
            if isinstance(item, dict):
                converted.append(
                    {
                        "src": item.get("src", ""),
                        "comment": item.get("comment", item.get("info", "")),
                    }
                )
            elif isinstance(item, str): # 兼容旧格式或纯字符串列表
                converted.append(
                    {
                        "src": item,
                        "comment": "",
                    }
                )
        self._set_table_data(converted)
        InfoBar.success(
            Localizer.get().local_glossary_completed,
            Localizer.get().text_preserve_loaded_do_not_translate_entries_settings.format(converted_count=len(converted)),
            parent=self,
        )

    def _save_to_config(self):
        entries = self._collect_table_data()
        self.config = Config().load()
        self.config.text_preserve_data = entries
        self.config.text_preserve_enable = True if entries else self.config.text_preserve_enable
        self.config.save()
        InfoBar.success(
            Localizer.get().local_glossary_saved,
            Localizer.get().text_preserve_saved_do_not_translate_entries_settings.format(entries_count=len(entries)),
            parent=self,
        )

    def _on_import_excel(self):
        if load_workbook is None:
            InfoBar.error(
                Localizer.get().error,
                Localizer.get().local_glossary_openpyxl_not_installed_so_excel_files_cannot,
                parent=self,
            )
            return

        path, _ = QFileDialog.getOpenFileName(
            self,
            Localizer.get().text_preserve_select_excel_file,
            "",
            Localizer.get().local_glossary_excel_files_xlsx
        )
        if not path:
            return
        try:
            workbook = load_workbook(path)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            header_map = self._build_header_map(headers)
            if "src" not in header_map:
                raise ValueError(Localizer.get().text_preserve_source_column_not_found_check_template)

            items: List[Dict[str, str]] = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                src = self._safe_cell(row, header_map.get("src"))
                comment = self._safe_cell(row, header_map.get("comment"))
                if not src:
                    continue
                items.append({"src": src, "comment": comment})

            self._set_table_data(items)
            InfoBar.success(
                Localizer.get().local_glossary_imported,
                Localizer.get().text_preserve_imported_do_not_translate_entries.format(items_count=len(items)),
                parent=self,
            )
        except Exception as e:
            self.logger.error(f"导入失败: {e}")
            InfoBar.error(
                Localizer.get().error,
                Localizer.get().extract_json_import_failed.format(e=e),
                parent=self,
            )

    def _on_export_excel(self):
        if Workbook is None:
            InfoBar.error(
                Localizer.get().error,
                Localizer.get().local_glossary_openpyxl_not_installed_so_excel_files_cannot_2,
                parent=self,
            )
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            Localizer.get().text_preserve_save_excel_file,
            "",
            Localizer.get().local_glossary_excel_files_xlsx
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        entries = self._collect_table_data()
        if not entries:
            InfoBar.warning(
                Localizer.get().notice,
                Localizer.get().local_glossary_table_empty_no_file_exported,
                parent=self,
            )
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "TextPreserve"
            sheet.append(list(self.HEADERS[:-1]))
            for item in entries:
                sheet.append([item.get("src", ""), item.get("comment", "")])
            workbook.save(path)
            InfoBar.success(
                Localizer.get().local_glossary_exported,
                Localizer.get().local_glossary_saved_2.format(path=path),
                parent=self,
            )
        except Exception as e:
            self.logger.error(f"导出失败: {e}")
            InfoBar.error(
                Localizer.get().error,
                Localizer.get().extract_json_export_failed.format(e=e),
                parent=self,
            )

    def _on_statistics_clicked(self) -> None:
        if self._statistics_worker and self._statistics_worker.isRunning():
            InfoBar.info(
                Localizer.get().notice,
                Localizer.get().local_glossary_hit_statistics_already_running,
                parent=self,
            )
            return

        entries = self._collect_table_data()
        if not entries:
            InfoBar.info(
                Localizer.get().notice,
                Localizer.get().text_preserve_there_no_do_not_translate_entries_analyze,
                parent=self,
            )
            return

        config = Config().load()
        self._statistics_snapshot_keys = [
            self._build_statistics_entry_key(entry) for entry in entries
        ]
        self._set_statistics_buttons_enabled(False)

        worker = RuleStatisticsWorker(
            mode = RuleStatisticsWorker.MODE_TEXT_PRESERVE,
            config = config,
            entries = entries,
            parent = self,
        )
        worker.finished.connect(self._on_statistics_finished)
        self._statistics_worker = worker

        worker.start()

    def _on_statistics_finished(self, success: bool, message: str, payload) -> None:
        self._set_statistics_buttons_enabled(True)

        worker = self._statistics_worker
        self._statistics_worker = None
        if worker is not None:
            worker.deleteLater()

        if success == False:
            error_message = (
                Localizer.get().rule_statistics_no_cached_entries
                if isinstance(payload, dict) and "cache_dir" in payload
                else Localizer.get().rule_statistics_unavailable
            )
            InfoBar.warning(
                Localizer.get().local_glossary_statistics_failed,
                error_message,
                parent=self,
            )
            return

        if not isinstance(payload, dict):
            InfoBar.warning(
                Localizer.get().local_glossary_statistics_failed,
                Localizer.get().local_glossary_statistics_result_has_invalid_format,
                parent=self,
            )
            return

        counts = payload.get("counts", [])
        if not isinstance(counts, list):
            InfoBar.warning(
                Localizer.get().local_glossary_statistics_failed,
                Localizer.get().local_glossary_statistics_result_does_not_contain_hit_counts,
                parent=self,
            )
            return

        current_entries = self._collect_table_data()
        current_keys = [self._build_statistics_entry_key(entry) for entry in current_entries]
        if current_keys != self._statistics_snapshot_keys:
            self._invalidate_statistics()
            InfoBar.warning(
                Localizer.get().notice,
                Localizer.get().text_preserve_entries_changed_run_statistics_again,
                parent=self,
            )
            return

        self.table.blockSignals(True)
        try:
            for row, count in enumerate(counts):
                if row >= self.table.rowCount():
                    break

                item = self.table.item(row, self.STATS_COLUMN)
                if item is None:
                    item = self._create_table_item("", editable=False)
                    self.table.setItem(row, self.STATS_COLUMN, item)
                item.setText(str(max(0, int(count))))
        finally:
            self.table.blockSignals(False)

        counted_item_total = int(payload.get("counted_item_total", 0))
        InfoBar.success(
            Localizer.get().local_glossary_statistics_completed,
            Localizer.get().text_preserve_analyzed_rules_across_cached_entries.format(counts_count=len(counts), counted_item_total=counted_item_total),
            parent=self,
        )

    # --- 工具方法 ---
    def _set_table_data(self, items: List[Dict[str, str]]):
        self.table.blockSignals(True)
        try:
            self.table.setRowCount(0)
            for item in items:
                row = self.table.rowCount()
                self.table.insertRow(row)
                self.table.setItem(row, 0, self._create_table_item(item.get("src", "")))
                self.table.setItem(row, 1, self._create_table_item(item.get("comment", "")))
                self.table.setItem(row, self.STATS_COLUMN, self._create_table_item("", editable=False))
        finally:
            self.table.blockSignals(False)
        self._invalidate_statistics()

    def _collect_table_data(self) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        rows = self.table.rowCount()
        for row in range(rows):
            src_item = self.table.item(row, 0)
            comment_item = self.table.item(row, 1)
            src = (src_item.text() if src_item else "").strip()
            comment = (comment_item.text() if comment_item else "").strip()
            if not src:
                continue
            results.append({"src": src, "comment": comment, "info": comment})
        return results

    @staticmethod
    def _normalize_src(text: str) -> str:
        if not text:
            return ""
        return text.strip().strip("\"'“”‘’").lower()

    @staticmethod
    def _merge_entries(base: Dict[str, str], incoming: Dict[str, str]) -> Dict[str, str]:
        def _clean(v: str) -> str:
            return v.strip() if isinstance(v, str) else ""

        merged = {"src": _clean(base.get("src")), "comment": _clean(base.get("comment"))}
        incoming_clean = {"src": _clean(incoming.get("src")), "comment": _clean(incoming.get("comment"))}

        # 保留有备注的
        if incoming_clean["comment"]:
            if not merged["comment"] or len(incoming_clean["comment"]) > len(merged["comment"]):
                merged["comment"] = incoming_clean["comment"]

        if incoming_clean["src"] and not merged["src"]:
            merged["src"] = incoming_clean["src"]
        return merged

    @staticmethod
    def _build_header_map(headers: List[str]) -> Dict[str, int]:
        alias = {
            "src": {"原文", "原始文本", "source", "src", "text"},
            "comment": {"备注", "说明", "comment", "note", "备注信息"},
        }
        mapping = {}
        for index, name in enumerate(headers):
            lower_name = name.lower()
            for key, options in alias.items():
                if lower_name in {opt.lower() for opt in options} and key not in mapping:
                    mapping[key] = index
        return mapping

    @staticmethod
    def _safe_cell(row, index: int) -> str:
        if index is None:
            return ""
        if index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()

    @staticmethod
    def _list_scan_candidates(config: Config) -> List[Path]:
        """根据当前配置推断变量扫描候选目录（按优先级排序并去重）"""
        raws = [
            getattr(config, "input_folder", ""),
            getattr(config, "output_folder", ""),
            getattr(config, "renpy_game_folder", ""),
        ]
        candidates: List[Path] = []
        for raw in raws:
            if not raw:
                continue
            path = Path(raw)
            if not path.exists():
                continue

            if path.is_file():
                if path.suffix.lower() == ".rpy":
                    candidates.append(path.parent)
                continue

            # 若选择的是 tl 目录，优先回退到上层 game 目录
            if path.name.lower() == "tl" and path.parent.exists():
                candidates.append(path.parent)

            game_child = path / "game"
            if game_child.exists() and game_child.is_dir():
                candidates.append(game_child)

            candidates.append(path)

        deduped: List[Path] = []
        seen: set[str] = set()
        for p in candidates:
            try:
                key = str(p.resolve()).lower()
            except Exception:
                key = str(p).lower()
            if key in seen:
                continue
            seen.add(key)
            deduped.append(p)
        return deduped

    @staticmethod
    def _count_rpy_files_without_tl(root: Path) -> int:
        """统计目录下可用于扫描变量的 rpy 数量（排除 tl 目录）"""
        count = 0
        for rpy_file in root.rglob("*.rpy"):
            if "tl" in [part.lower() for part in rpy_file.parts]:
                continue
            count += 1
        return count

    def _on_rescan_variables(self):
        """重新扫描游戏目录，提取[variable]变量引用到禁翻表（清空旧数据）"""
        import re
        
        # 重新加载配置以获取最新目录配置
        self.config = Config().load()

        candidates = self._list_scan_candidates(self.config)
        if not candidates:
            InfoBar.warning(
                Localizer.get().warning,
                Localizer.get().text_preserve_no_folder_available_scan_set_input_output,
                parent=self,
            )
            return

        # 按候选优先级选择第一个可扫描目录：
        # input/output 优先于 renpy_game_folder，避免跳到历史项目。
        game_path = None
        fallback_path = None
        for candidate in candidates:
            if fallback_path is None:
                fallback_path = candidate
            try:
                count = self._count_rpy_files_without_tl(candidate)
            except Exception:
                count = -1
            if count > 0:
                game_path = candidate
                break

        if game_path is None:
            game_path = fallback_path

        if game_path is None or not game_path.exists():
            InfoBar.error(
                Localizer.get().error,
                Localizer.get().text_preserve_could_not_determine_which_folder_scan,
                parent=self,
            )
            return

        # 正则匹配 [variable_name]
        RE_VARIABLE_IN_TEXT = re.compile(r'\[([\w.]+)\]')
        
        found_preserves = set()
        try:
            for rpy_file in game_path.rglob("*.rpy"):
                # 跳过 tl 目录，避免将翻译产物中的占位污染禁翻表
                if "tl" in [part.lower() for part in rpy_file.parts]:
                    continue
                try:
                    content = rpy_file.read_text(encoding="utf-8", errors="ignore")
                    var_matches = RE_VARIABLE_IN_TEXT.findall(content)
                    for var_name in var_matches:
                        found_preserves.add(f"[{var_name}]")
                except Exception:
                    pass
        except Exception as e:
            InfoBar.error(
                Localizer.get().error,
                Localizer.get().text_preserve_scan_failed.format(e=e),
                parent=self,
            )
            return
        
        if not found_preserves:
            # 清空禁翻表
            self.config.text_preserve_data = []
            self.config.text_preserve_enable = False
            self.config.save()
            self._load_from_config()
            InfoBar.info(
                Localizer.get().notice,
                Localizer.get().text_preserve_no_variable_references_found_list_cleared_scanned.format(game_path=game_path),
                parent=self,
            )
            return
        
        # 完全清空旧数据，只保留新扫描的 [variable]
        new_preserves = []
        for text in sorted(found_preserves):
            new_preserves.append({"src": text})
        
        # 保存到配置
        self.config.text_preserve_data = new_preserves
        self.config.text_preserve_enable = True
        self.config.save()
        
        # 刷新表格
        self._load_from_config()
        
        InfoBar.success(
            Localizer.get().local_glossary_completed,
            Localizer.get().text_preserve_found_variable_references_scanned.format(new_preserves_count=len(new_preserves), game_path=game_path),
            parent=self,
        )
