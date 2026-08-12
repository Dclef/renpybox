"""
HTML 导入 / 转换工具
支持从 HTML 提取文本，以及将 TXT 转为可用于翻译对照的 HTML
"""

import os
from typing import List

from bs4 import BeautifulSoup
from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QFileDialog,
)
from qfluentwidgets import (
    BodyLabel,
    CardWidget,
    PushButton,
    PrimaryPushButton,
    LineEdit,
    CheckBox,
    ComboBox,
    InfoBar,
    FluentIcon,
    StrongBodyLabel,
    TitleLabel,
)

from base.Base import Base
from base.LogManager import LogManager
from module.Localizer.Localizer import Localizer


class HtmlImportPage(Base, QWidget):
    """HTML 导入工具"""

    def __init__(self, object_name: str, parent=None):
        Base.__init__(self)
        QWidget.__init__(self, parent)
        self.setObjectName(object_name)
        self.logger = LogManager.get()

        self._init_ui()

    # --- UI ---
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = TitleLabel(
            Localizer.localize("HTML 导入 / 转换", "HTML Import / Conversion"), self
        )
        layout.addWidget(title)

        desc = BodyLabel(
            Localizer.localize(
                "支持从 HTML 文件提取翻译文本，或将 TXT 转换为 HTML，方便在浏览器或翻译工具中审阅。",
                "Extract translation text from HTML files, or convert TXT files to HTML for "
                "review in a browser or translation tool.",
            ),
            self,
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)

        layout.addWidget(self._build_html_to_txt_card())
        layout.addWidget(self._build_txt_to_html_card())
        layout.addWidget(self._build_excel_to_txt_card())
        layout.addStretch(1)

    def _build_html_to_txt_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 16)
        v_layout.setSpacing(10)

        title = StrongBodyLabel(Localizer.localize("HTML → TXT", "HTML → TXT"), self)
        v_layout.addWidget(title)

        # 输入
        input_row = QHBoxLayout()
        input_row.addWidget(QLabel(Localizer.localize("HTML 文件:", "HTML File:")))
        self.html_input_edit = LineEdit()
        self.html_input_edit.setPlaceholderText(
            Localizer.localize(
                "选择包含 <h6> 节点的 HTML 文件",
                "Select an HTML file containing <h6> elements",
            )
        )
        input_row.addWidget(self.html_input_edit, 1)
        btn_browse_html = PushButton(Localizer.get().browse, icon=FluentIcon.FOLDER)
        btn_browse_html.clicked.connect(self._browse_html_file)
        input_row.addWidget(btn_browse_html)
        v_layout.addLayout(input_row)

        # 输出
        output_row = QHBoxLayout()
        output_row.addWidget(QLabel(Localizer.localize("输出 TXT:", "Output TXT:")))
        self.html_output_edit = LineEdit()
        self.html_output_edit.setPlaceholderText(
            Localizer.localize(
                "默认与输入同名，可留空",
                "Leave blank to use the input file name",
            )
        )
        output_row.addWidget(self.html_output_edit, 1)
        btn_browse_txt = PushButton(Localizer.get().browse, icon=FluentIcon.FOLDER)
        btn_browse_txt.clicked.connect(self._browse_txt_output)
        output_row.addWidget(btn_browse_txt)
        v_layout.addLayout(output_row)

        convert_btn = PrimaryPushButton(
            Localizer.localize("执行导出", "Export"), icon=FluentIcon.DOWNLOAD
        )
        convert_btn.clicked.connect(self._convert_html_to_txt)
        v_layout.addWidget(convert_btn)

        return card

    def _build_excel_to_txt_card(self) -> CardWidget:
        """Excel → TXT"""
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 16)
        v_layout.setSpacing(10)

        title = StrongBodyLabel(Localizer.localize("Excel → TXT", "Excel → TXT"), self)
        v_layout.addWidget(title)

        # 选择 Excel
        row1 = QHBoxLayout()
        row1.addWidget(QLabel(Localizer.localize("Excel 文件:", "Excel File:")))
        self.excel_input_edit = LineEdit()
        self.excel_input_edit.setPlaceholderText(
            Localizer.localize(
                "选择由本工具导出的 Excel（包含 原文/译文 列）",
                "Select an Excel file exported by this tool (with Original/Translation columns)",
            )
        )
        row1.addWidget(self.excel_input_edit, 1)
        btn_browse_excel = PushButton(Localizer.get().browse, icon=FluentIcon.FOLDER)
        btn_browse_excel.clicked.connect(self._browse_excel_input)
        row1.addWidget(btn_browse_excel)
        v_layout.addLayout(row1)

        # 输出 TXT
        row2 = QHBoxLayout()
        row2.addWidget(QLabel(Localizer.localize("输出 TXT:", "Output TXT:")))
        self.excel_txt_output_edit = LineEdit()
        self.excel_txt_output_edit.setPlaceholderText(
            Localizer.localize(
                "默认与 Excel 同名，可留空",
                "Leave blank to use the Excel file name",
            )
        )
        row2.addWidget(self.excel_txt_output_edit, 1)
        btn_browse_excel_txt = PushButton(Localizer.get().browse, icon=FluentIcon.FOLDER)
        btn_browse_excel_txt.clicked.connect(self._browse_excel_txt_output)
        row2.addWidget(btn_browse_excel_txt)
        v_layout.addLayout(row2)

        # 选择导出列
        row3 = QHBoxLayout()
        row3.addWidget(QLabel(Localizer.localize("导出列:", "Export Column:")))
        self.excel_column_combo = ComboBox()
        self.excel_column_combo.addItem(
            Localizer.localize("译文", "Translation"), userData="译文"
        )
        self.excel_column_combo.addItem(
            Localizer.localize("原文", "Original"), userData="原文"
        )
        row3.addWidget(self.excel_column_combo, 1)
        v_layout.addLayout(row3)

        export_btn = PrimaryPushButton(
            Localizer.localize("导出 TXT", "Export TXT"), icon=FluentIcon.SAVE
        )
        export_btn.clicked.connect(self._convert_excel_to_txt)
        v_layout.addWidget(export_btn)

        return card

    def _build_txt_to_html_card(self) -> CardWidget:
        card = CardWidget(self)
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 12, 16, 16)
        v_layout.setSpacing(10)

        title = StrongBodyLabel(Localizer.localize("TXT → HTML", "TXT → HTML"), self)
        v_layout.addWidget(title)

        # 输入
        txt_row = QHBoxLayout()
        txt_row.addWidget(QLabel(Localizer.localize("TXT 文件:", "TXT File:")))
        self.txt_input_edit = LineEdit()
        self.txt_input_edit.setPlaceholderText(
            Localizer.localize(
                "每行一条文本，将转换为 <h6> 节点",
                "Each line becomes an <h6> element",
            )
        )
        txt_row.addWidget(self.txt_input_edit, 1)
        btn_browse_txt_in = PushButton(Localizer.get().browse, icon=FluentIcon.FOLDER)
        btn_browse_txt_in.clicked.connect(self._browse_txt_input)
        txt_row.addWidget(btn_browse_txt_in)
        v_layout.addLayout(txt_row)

        # 输出
        html_row = QHBoxLayout()
        html_row.addWidget(QLabel(Localizer.localize("输出 HTML:", "Output HTML:")))
        self.txt_output_edit = LineEdit()
        self.txt_output_edit.setPlaceholderText(
            Localizer.localize(
                "默认与输入同名，可留空",
                "Leave blank to use the input file name",
            )
        )
        html_row.addWidget(self.txt_output_edit, 1)
        btn_browse_html_out = PushButton(Localizer.get().browse, icon=FluentIcon.FOLDER)
        btn_browse_html_out.clicked.connect(self._browse_html_output)
        html_row.addWidget(btn_browse_html_out)
        v_layout.addLayout(html_row)

        self.wrap_data_check = CheckBox(
            Localizer.localize(
                "写入附加数据（保留原文/译文结构）",
                "Include additional data (preserve the original/translation structure)",
            )
        )
        v_layout.addWidget(self.wrap_data_check)

        convert_btn = PrimaryPushButton(
            Localizer.localize("生成 HTML", "Generate HTML"), icon=FluentIcon.UP
        )
        convert_btn.clicked.connect(self._convert_txt_to_html)
        v_layout.addWidget(convert_btn)

        return card

    # --- 槽函数 ---
    def _browse_html_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            Localizer.localize("选择 HTML 文件", "Select HTML File"),
            "",
            Localizer.localize(
                "HTML 文件 (*.html *.htm)", "HTML Files (*.html *.htm)"
            ),
        )
        if path:
            self.html_input_edit.setText(path)

    def _browse_txt_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            Localizer.localize("选择输出 TXT 路径", "Select Output TXT Path"),
            "",
            Localizer.localize("文本文件 (*.txt)", "Text Files (*.txt)"),
        )
        if path:
            if not path.lower().endswith(".txt"):
                path += ".txt"
            self.html_output_edit.setText(path)

    def _browse_txt_input(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            Localizer.localize("选择 TXT 文件", "Select TXT File"),
            "",
            Localizer.localize("文本文件 (*.txt)", "Text Files (*.txt)"),
        )
        if path:
            self.txt_input_edit.setText(path)

    def _browse_html_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            Localizer.localize("选择输出 HTML 路径", "Select Output HTML Path"),
            "",
            Localizer.localize(
                "HTML 文件 (*.html *.htm)", "HTML Files (*.html *.htm)"
            ),
        )
        if path:
            if not path.lower().endswith((".html", ".htm")):
                path += ".html"
            self.txt_output_edit.setText(path)

    def _browse_excel_input(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            Localizer.localize("选择 Excel 文件", "Select Excel File"),
            "",
            Localizer.localize("Excel 文件 (*.xlsx)", "Excel Files (*.xlsx)"),
        )
        if path:
            self.excel_input_edit.setText(path)

    def _browse_excel_txt_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            Localizer.localize("选择输出 TXT 路径", "Select Output TXT Path"),
            "",
            Localizer.localize("文本文件 (*.txt)", "Text Files (*.txt)"),
        )
        if path:
            if not path.lower().endswith(".txt"):
                path += ".txt"
            self.excel_txt_output_edit.setText(path)

    def _convert_html_to_txt(self):
        html_path = self.html_input_edit.text().strip()
        if not html_path:
            InfoBar.warning(
                Localizer.get().notice,
                Localizer.localize("请先选择 HTML 文件", "Select an HTML file first."),
                parent=self,
            )
            return
        if not os.path.isfile(html_path):
            InfoBar.error(
                Localizer.get().error,
                Localizer.localize("HTML 文件不存在", "The HTML file does not exist."),
                parent=self,
            )
            return

        output_path = self.html_output_edit.text().strip()
        if not output_path:
            root, _ = os.path.splitext(html_path)
            output_path = root + ".txt"

        try:
            strings = self._read_html_strings(html_path)
            if not strings:
                raise ValueError(
                    Localizer.localize(
                        "未在 HTML 中找到 <h6> 节点，请确认文件格式。",
                        "No <h6> elements were found in the HTML file. Check the file format.",
                    )
                )
            with open(output_path, "w", encoding="utf-8") as writer:
                writer.write("\n".join(strings))
            InfoBar.success(
                Localizer.get().complete,
                Localizer.localize(
                    "已导出到 {output_path}", "Exported to {output_path}"
                ).format(output_path=output_path),
                parent=self,
            )
        except Exception as e:
            self.logger.error(f"HTML 导出失败: {e}")
            InfoBar.error(
                Localizer.get().error,
                Localizer.localize("导出失败: {e}", "Export failed: {e}").format(e=e),
                parent=self,
            )

    def _convert_txt_to_html(self):
        txt_path = self.txt_input_edit.text().strip()
        if not txt_path:
            InfoBar.warning(
                Localizer.get().notice,
                Localizer.localize("请先选择 TXT 文件", "Select a TXT file first."),
                parent=self,
            )
            return
        if not os.path.isfile(txt_path):
            InfoBar.error(
                Localizer.get().error,
                Localizer.localize("TXT 文件不存在", "The TXT file does not exist."),
                parent=self,
            )
            return

        output_path = self.txt_output_edit.text().strip()
        if not output_path:
            root, _ = os.path.splitext(txt_path)
            output_path = root + ".html"

        try:
            with open(txt_path, "r", encoding="utf-8") as reader:
                lines = [line.rstrip("\n") for line in reader]
            if not lines:
                raise ValueError(
                    Localizer.localize("TXT 文件为空", "The TXT file is empty.")
                )

            html = self._build_html_content(lines, keep_data=self.wrap_data_check.isChecked())
            with open(output_path, "w", encoding="utf-8") as writer:
                writer.write(html)
            InfoBar.success(
                Localizer.get().complete,
                Localizer.localize(
                    "已生成 HTML：{output_path}", "HTML generated: {output_path}"
                ).format(output_path=output_path),
                parent=self,
            )
        except Exception as e:
            self.logger.error(f"TXT 转 HTML 失败: {e}")
            InfoBar.error(
                Localizer.get().error,
                Localizer.localize("生成失败: {e}", "Generation failed: {e}").format(e=e),
                parent=self,
            )

    def _convert_excel_to_txt(self):
        try:
            from openpyxl import load_workbook
        except Exception:
            InfoBar.error(
                Localizer.get().error,
                Localizer.localize(
                    "未安装 openpyxl，无法读取 Excel",
                    "openpyxl is not installed, so Excel files cannot be read.",
                ),
                parent=self,
            )
            return

        excel_path = self.excel_input_edit.text().strip()
        if not excel_path:
            InfoBar.warning(
                Localizer.get().notice,
                Localizer.localize("请先选择 Excel 文件", "Select an Excel file first."),
                parent=self,
            )
            return
        if not os.path.isfile(excel_path):
            InfoBar.error(
                Localizer.get().error,
                Localizer.localize("Excel 文件不存在", "The Excel file does not exist."),
                parent=self,
            )
            return

        output_path = self.excel_txt_output_edit.text().strip()
        if not output_path:
            root, _ = os.path.splitext(excel_path)
            output_path = root + ".txt"

        target_col_name = self.excel_column_combo.currentData()

        try:
            book = load_workbook(excel_path)
            lines: list[str] = []
            for sheet_name in book.sheetnames:
                if sheet_name in ("元数据", "Metadata"):
                    continue
                sheet = book[sheet_name]
                if sheet.max_row == 0 or sheet.max_column == 0:
                    continue
                headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
                # 兼容别名
                alias_map = {
                    "原文": {
                        "原文",
                        "original",
                        "原文（勿修改此列）",
                        "Source (Do Not Edit)",
                    },
                    "译文": {
                        "译文",
                        "translation",
                        "译文（勿修改此列）",
                        "Translation (Do Not Edit)",
                    },
                }
                col_index = None
                for idx, name in enumerate(headers):
                    if name.lower() in {v.lower() for v in alias_map.get(target_col_name, {target_col_name})}:
                        col_index = idx
                        break
                if col_index is None:
                    continue
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    val = row[col_index] if col_index < len(row) else None
                    if val is None:
                        continue
                    text = str(val).strip()
                    if text:
                        lines.append(text)

            if not lines:
                InfoBar.warning(
                    Localizer.get().notice,
                    Localizer.localize(
                        "未在 Excel 中找到可导出的内容",
                        "No exportable content was found in the Excel file.",
                    ),
                    parent=self,
                )
                return

            with open(output_path, "w", encoding="utf-8") as writer:
                writer.write("\n".join(lines))
            InfoBar.success(
                Localizer.get().complete,
                Localizer.localize(
                    "已导出 {count} 行到 {output_path}",
                    "Exported {count} row(s) to {output_path}",
                ).format(count=len(lines), output_path=output_path),
                parent=self,
            )
        except Exception as e:
            LogManager.get().error(f"Excel → TXT 导出失败: {e}")
            InfoBar.error(
                Localizer.get().error,
                Localizer.localize("导出失败: {e}", "Export failed: {e}").format(e=e),
                parent=self,
            )

    # --- 工具函数 ---
    @staticmethod
    def _read_html_strings(path: str) -> List[str]:
        with open(path, "r", encoding="utf-8") as reader:
            soup = BeautifulSoup(reader, "html.parser")
        strings = [tag.get_text() for tag in soup.find_all("h6")]
        if not strings:  # 兼容其它标签
            strings = [tag.get_text() for tag in soup.find_all(["p", "div"])]
        return [s.replace("\r", "").strip() for s in strings if s and s.strip()]

    @staticmethod
    def _build_html_content(lines: List[str], keep_data: bool) -> str:
        soup = BeautifulSoup("<html><head><meta charset='utf-8'></head><body></body></html>", "html.parser")
        body = soup.body

        data_payload = []
        for idx, text in enumerate(lines):
            text = text or ""
            h6 = soup.new_tag("h6")
            h6.string = text
            body.append(h6)
            if keep_data:
                data_payload.append(
                    {
                        "line": idx,
                        "original": text,
                        "target": text,
                        "current": text,
                    }
                )

        if keep_data and data_payload:
            data_div = soup.new_tag("div", id="data", style="display: none;")
            data_div.string = str(data_payload)
            body.append(data_div)

        return str(soup)
